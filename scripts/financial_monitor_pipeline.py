import argparse
import calendar
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
import hashlib
import json
import os
import re
import signal
import threading
import time
import unicodedata
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


WORKSPACE = Path(__file__).resolve().parents[1]
DEFAULT_CLIENTS_ROOT = Path(
    "/Users/syscap/Library/CloudStorage/GoogleDrive-rbarron@syscap.com.mx/Shared drives/"
    "Axcess - Crédito y Riesgo/1. Clientes/2. Activos & Prospectos"
)
MAPPING_PATH = WORKSPACE / "config/account_mapping_memory.tsv"
CLIENT_METADATA_PATH = WORKSPACE / "config/client_metadata_template.tsv"
CLIENT_FOLLOWUPS_PATH = WORKSPACE / "config/client_followups.tsv"
OUTPUT_DIR = WORKSPACE / "outputs/financial_monitor"
CACHE_DIR = WORKSPACE / "outputs/.cache/pdf_accounts"
AUXILIARY_SOURCE_ROOTS = [
    Path(
        "/Users/syscap/Library/CloudStorage/GoogleDrive-rbarron@syscap.com.mx/Shared drives/"
        "Axcess - Crédito y Riesgo/4. Proyectos Estratégicos/"
        "Automatización Estados Financieros - Denise/Documentos de avances"
    )
]
AMOUNT_RE = r"\(?-?\s*(?:\d+\s+)?\d{1,3}(?:\s*,\s*\d{3})+(?:\.\d+)?\)?|\(?-?\s*\d+(?:\.\d+)?\)?"
PAGE_TIMEOUT_SECONDS = 18
SUPPORTED_SUFFIXES = {".pdf", ".xlsx", ".xls", ".webloc"}
EXTRACTABLE_SUFFIXES = {".pdf", ".xlsx", ".xls"}
FRONT_SHEETS = {"Inicio", "CRM Clientes", "Loan Tape Cliente", "Estados Financieros", "Actualizar"}
HIDDEN_SHEETS = {"Conceptos", "Cuentas Extraidas", "Credit Evidence", "Golden Sample", "Mapping Memory"}
FILTERABLE_FRONT_SHEETS = {"CRM Clientes", "Loan Tape Cliente", "Estados Financieros", "Actualizar"}
FOLLOWUP_COLUMNS = ["responsable", "proxima_accion", "fecha_actualizacion", "seguimiento_notas"]
FACILITY_COLUMNS = ["facility_id", "facility_name", "facility_covenants"]
FINANCIAL_STATEMENT_LAYOUT = [
    ("BG", "Activo", "Efectivo e inversiones", "efectivo_inversiones"),
    ("BG", "Activo", "Clientes arrendamiento", "clientes_arrendamiento"),
    ("BG", "Activo", "Clientes factoraje", "clientes_factoraje"),
    ("BG", "Activo", "Estimacion preventiva", "estimacion_preventiva"),
    ("BG", "Activo", "Cartera bruta", "cartera_bruta"),
    ("BG", "Activo", "Cartera neta credito", "cartera_neta_credito"),
    ("BG", "Activo", "Otros activos generadores", "otros_activos_generadores"),
    ("BG", "Activo", "Total activo", "total_activo"),
    ("BG", "Pasivo", "Fondeadores CP", "fondeadores_cp"),
    ("BG", "Pasivo", "Fondeadores LP", "fondeadores_lp"),
    ("BG", "Pasivo", "Deuda CP proxy", "deuda_cp_proxy"),
    ("BG", "Pasivo", "Deuda LP proxy", "deuda_lp_proxy"),
    ("BG", "Pasivo", "Total pasivo", "total_pasivo"),
    ("BG", "Capital", "Total capital contable", "total_capital_contable"),
    ("ER", "Ingresos", "Ingresos totales", "ingresos_totales"),
    ("ER", "Ingresos", "Ingresos intereses", "ingresos_intereses"),
    ("ER", "Ingresos", "Ingresos por comisiones", "ingresos_por_comisiones"),
    ("ER", "Costos", "Costo de ventas", "costo_ventas"),
    ("ER", "Resultado", "Utilidad bruta", "utilidad_bruta"),
    ("ER", "Gastos", "Gastos de operacion", "gastos_operacion"),
    ("ER", "Gastos", "Gastos financieros", "gastos_financieros"),
    ("ER", "Ingresos", "Productos financieros", "productos_financieros"),
    ("ER", "Ingresos", "Intereses a favor", "intereses_a_favor"),
    ("ER", "Resultado", "Utilidad operacion", "utilidad_operacion"),
    ("ER", "Resultado", "Utilidad neta", "utilidad_neta"),
]
FIN_MONITOR_PROCESSES = [
    {
        "process": "Dashboard Comercial",
        "line": "Comercial",
        "includes": "KPIs, pipeline, cartera y vista ejecutiva",
        "app_view": "Inicio",
        "objective": "Ver desempeno comercial y alertas principales.",
    },
    {
        "process": "CRM Comercial",
        "line": "Comercial",
        "includes": "Clientes, responsables, proxima accion y seguimiento",
        "app_view": "CRM Clientes",
        "objective": "Gestionar relacion comercial y ciclo de atencion.",
    },
    {
        "process": "Monitoreo",
        "line": "Linea de vida",
        "includes": "Linea de vida, clientes, alertas y bloqueo principal",
        "app_view": "Inicio / CRM Clientes",
        "objective": "Vigilar estado del cliente y eventos de riesgo.",
    },
    {
        "process": "Linea de Analisis",
        "line": "Analisis",
        "includes": "Benchmarking, consolidacion y Z Core / Z-Score",
        "app_view": "Razones / QA",
        "objective": "Comparar, consolidar y medir riesgo financiero.",
    },
    {
        "process": "MI Quality",
        "line": "Data Quality",
        "includes": "Ingestion, validacion y consolidacion",
        "app_view": "Actualizar / Documentos / Auditoria",
        "objective": "Asegurar datos confiables, trazables y auditables.",
    },
]
MONTHS = {
    "ene": "01",
    "enero": "01",
    "feb": "02",
    "febrero": "02",
    "mar": "03",
    "marzo": "03",
    "abr": "04",
    "abril": "04",
    "may": "05",
    "mayo": "05",
    "jun": "06",
    "junio": "06",
    "jul": "07",
    "julio": "07",
    "ago": "08",
    "agosto": "08",
    "sep": "09",
    "sept": "09",
    "septiembre": "09",
    "oct": "10",
    "octubre": "10",
    "nov": "11",
    "noviembre": "11",
    "dic": "12",
    "diciembre": "12",
}
REQUIRED_TOTAL_LABELS = [
    "TOTAL ACTIVO",
    "TOTAL PASIVO",
    "TOTAL CAPITAL CONTABLE",
    "TOTAL INGRESOS",
    "TOTAL COSTO DE VENTAS",
    "UTLIDAD (PÉRDIDA) BRUTA",
    "UTILIDAD (PÉRDIDA) BRUTA",
    "TOTAL GASTOS DE OPERACIÓN",
    "UTILIDAD (PÉRDIDA) EN OPERACIÓN",
    "UTLIDAD (PÉRDIDA) EN OPERACIÓN",
    "UTILIDAD (PÉRDIDA) NETA",
]


class PageTimeout(Exception):
    pass


def timeout_handler(signum, frame):
    raise PageTimeout(f"page extraction timed out after {PAGE_TIMEOUT_SECONDS}s")


def _can_use_signal_alarm():
    return threading.current_thread() is threading.main_thread()


def normalize(text):
    text = "" if text is None else str(text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_number(value):
    if value is None:
        return None
    value = str(value).replace(" ", "")
    if value in {"", "-"}:
        return 0.0
    negative = value.startswith("(") and value.endswith(")")
    value = value.strip("()").replace(",", "")
    try:
        number = float(value)
    except ValueError:
        return None
    return -number if negative else number


def money_candidates(text):
    text = "" if text is None else str(text)
    text = re.sub(r"-?\d+(?:\.\d+)?%", " ", text)
    return re.findall(AMOUNT_RE, text)


def first_money_after_label(line, label):
    match = re.search(re.escape(label), line, re.I)
    if not match:
        return None
    candidates = money_candidates(line[match.end() :])
    return clean_number(candidates[0]) if candidates else None


def period_from_name(path):
    match = re.search(r"(\d{6})", path.name)
    if not match:
        name = normalize(path.name)
        month_match = re.search(
            r"\b(ene|enero|feb|febrero|mar|marzo|abr|abril|may|mayo|jun|junio|jul|julio|ago|agosto|sep|sept|septiembre|oct|octubre|nov|noviembre|dic|diciembre)\b\s*(20\d{2})\b",
            name,
        )
        if month_match:
            return f"{month_match.group(2)}-{MONTHS[month_match.group(1)]}-01"
        return ""
    token = match.group(1)
    return f"20{token[:2]}-{token[2:4]}-{token[4:6]}"


def month_end_period(year, month):
    day = calendar.monthrange(int(year), int(month))[1]
    return f"{int(year):04d}-{int(month):02d}-{day:02d}"


def period_from_label(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if hasattr(value, "year") and hasattr(value, "month"):
        return month_end_period(value.year, value.month)
    text = str(value).strip()
    if not text:
        return ""
    match = re.search(r"\b(20\d{2})[-/](\d{1,2})(?:[-/]\d{1,2})?\b", text)
    if match:
        return month_end_period(match.group(1), match.group(2))
    normalized = normalize(text)
    month_match = re.search(
        r"\b(ene|enero|feb|febrero|mar|marzo|abr|abril|may|mayo|jun|junio|jul|julio|ago|agosto|sep|sept|septiembre|oct|octubre|nov|noviembre|dic|diciembre)\b\s*(20\d{2})\b",
        normalized,
    )
    if month_match:
        return month_end_period(month_match.group(2), MONTHS[month_match.group(1)])
    return ""


def classify_statement(path):
    name = normalize(path.name)
    if re.search(r"\bbg\b|balance|situacion financiera", name):
        return "BG"
    if re.search(r"\ber\b|estado de resultados|resultados", name):
        return "ER"
    if re.search(r"\bef\b|eeff|estados financieros", name):
        return "EF"
    return "UNKNOWN"


def classify_statement_name(value):
    name = normalize(value)
    if re.search(r"\bbg\b|balance|situacion financiera", name):
        return "BG"
    if re.search(r"\ber\b|estado de resultados|resultados", name):
        return "ER"
    if re.search(r"\bflujo\b|cash flow", name):
        return "CF"
    return "UNKNOWN"


def classify_source_quality(path):
    name = normalize(path.name)
    suffix = path.suffix.lower()
    if "modelo" in name or "proyeccion" in name or "proyecciones" in name:
        return "projection_or_model"
    if "sat" in name or "declaracion" in name or "declaracion anual" in name:
        return "tax_statement"
    if "dictamin" in name or "firmado" in name:
        return "signed_or_audited"
    if suffix in {".xlsx", ".xls"}:
        return "spreadsheet"
    if suffix == ".webloc":
        return "drive_link"
    if suffix == ".pdf":
        return "pdf_text_or_table"
    return "unknown"


def infer_statement_frequency(path):
    name = normalize(path.name)
    if "mensual" in name:
        return "Mensual"
    if "acumulado" in name or re.search(r"\bef\b|eeff|estados financieros", name):
        return "Mensual Acumulado"
    if re.search(r"dictamin|anual|sat|declaracion", name):
        return "Anual"
    return ""


def infer_unit_scale(path, text=""):
    haystack = normalize(f"{path.name} {text[:1000]}")
    if re.search(r"miles de pesos|cifras en miles|en miles", haystack):
        return 1000
    if re.search(r"millones de pesos|cifras en millones|en millones", haystack):
        return 1000000
    return 1


def log_step(label, start=None):
    if start is None:
        print(f"[pipeline] {label}", flush=True)
        return time.perf_counter()
    elapsed = time.perf_counter() - start
    print(f"[pipeline] {label} ({elapsed:.1f}s)", flush=True)
    return elapsed


def cache_path_for(path, max_pages, max_file_mb):
    resolved = str(Path(path).expanduser())
    key = hashlib.sha1(f"{resolved}|pages={max_pages}|mb={max_file_mb}".encode("utf-8")).hexdigest()
    return CACHE_DIR / f"{key}.json"


def file_fingerprint(path):
    stat = Path(path).stat()
    return {"path": str(path), "size": stat.st_size, "mtime_ns": stat.st_mtime_ns}


def read_pdf_cache(path, max_pages, max_file_mb):
    cache_path = cache_path_for(path, max_pages, max_file_mb)
    if not cache_path.exists():
        return None
    try:
        payload = json.loads(cache_path.read_text())
        if payload.get("fingerprint") == file_fingerprint(path):
            return payload
    except (OSError, json.JSONDecodeError):
        return None
    return None


def write_pdf_cache(path, max_pages, max_file_mb, rows, error):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "fingerprint": file_fingerprint(path),
        "rows": rows,
        "error": error,
    }
    cache_path_for(path, max_pages, max_file_mb).write_text(json.dumps(payload, ensure_ascii=False))


def client_financial_dir(clients_root, client):
    return clients_root / client / "1. Data Room/3. Información Financiera/1. Estados Financieros"


def client_root_dir(clients_root, client):
    direct = clients_root / client
    if direct.exists():
        return direct
    if not clients_root.exists():
        return direct
    target = normalize(client)
    for candidate in clients_root.iterdir():
        if candidate.is_dir() and normalize(candidate.name) == target:
            return candidate
    return direct


def find_financial_dirs(clients_root, client):
    root = client_root_dir(clients_root, client)
    direct = root / "1. Data Room/3. Información Financiera/1. Estados Financieros"
    dirs = []
    if direct.exists():
        dirs.append(direct)
    if root.exists():
        has_flat_financial_files = any(path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES for path in root.iterdir())
        if has_flat_financial_files and root not in dirs:
            dirs.append(root)
        for path in root.rglob("1. Estados Financieros"):
            if "3. Información Financiera" in str(path) and path not in dirs:
                dirs.append(path)
    return dirs


def scan_credit_line_evidence(clients_root, clients):
    rows = []
    transaction_markers = re.compile(r"(transacciones|contratos|term sheets|monitoreo)", re.I)
    credit_markers = re.compile(r"(contrato|car[aá]tula|pagare|pagar[eé]|cesi[oó]n|term sheet)", re.I)
    high_signal = re.compile(r"(contrato|car[aá]tula|pagare|pagar[eé]|term sheet)", re.I)
    amount_markers = re.compile(r"(\$|m\b|mxn|usd|cp|lp)", re.I)
    allowed_suffixes = {".pdf", ".docx", ".xlsx", ".xls"}
    for client in clients:
        root = client_root_dir(clients_root, client)
        if not root.exists():
            rows.append({"client": client, "credit_line_found": False, "evidence_type": "missing_client_dir", "path": str(root), "filename": "", "confidence": 0.0})
            continue
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            if path.suffix.lower() not in allowed_suffixes:
                continue
            path_text = str(path)
            if not transaction_markers.search(path_text):
                continue
            name = path.name
            if not high_signal.search(name):
                continue
            if not credit_markers.search(name) and not credit_markers.search(path_text):
                continue
            confidence = 0.65
            if "3. Contratos" in path_text:
                confidence += 0.2
            if amount_markers.search(path_text):
                confidence += 0.1
            rows.append(
                {
                    "client": client,
                    "credit_line_found": True,
                    "evidence_type": "drive_transaction_file",
                    "path": path_text,
                    "filename": name,
                    "confidence": min(confidence, 0.95),
                }
            )
    return pd.DataFrame(rows)


def load_client_metadata(path):
    if not path or not Path(path).exists():
        return pd.DataFrame()
    if str(path).lower().endswith((".xlsx", ".xls")):
        metadata = pd.read_excel(path)
    else:
        metadata = pd.read_csv(path, sep="\t")
    if "client" not in metadata.columns:
        raise ValueError("Client metadata must include a 'client' column.")
    rename_map = {}
    for col in metadata.columns:
        key = normalize(col)
        if key in {"facility", "facility id", "id facility", "linea credito id", "linea de credito id"}:
            rename_map[col] = "facility_id"
        elif key in {"facility name", "facility nombre", "nombre facility", "linea credito", "linea de credito"}:
            rename_map[col] = "facility_name"
        elif key in {"facility covenants", "covenants facility", "covenants", "covenants seleccionados"}:
            rename_map[col] = "facility_covenants"
    if rename_map:
        metadata = metadata.rename(columns=rename_map)
    metadata["client"] = metadata["client"].astype(str)
    return metadata


def _metadata_value_matches(value, target):
    normalized = normalize(value)
    if normalized == target:
        return True
    parts = [normalize(part) for part in re.split(r"[|,;/]+", "" if value is None else str(value))]
    return target in parts


def filter_client_metadata_by_facility(metadata, facility):
    if metadata.empty or not facility:
        return metadata
    facility_cols = [col for col in FACILITY_COLUMNS if col in metadata.columns]
    if not facility_cols:
        raise ValueError(
            "Facility filter requested, but client metadata has no facility columns. "
            "Add facility_id, facility_name or facility_covenants."
        )
    target = normalize(facility)
    mask = pd.Series(False, index=metadata.index)
    for col in facility_cols:
        mask = mask | metadata[col].map(lambda value: _metadata_value_matches(value, target))
    filtered = metadata[mask].copy()
    if filtered.empty:
        available = sorted(
            {
                str(value).strip()
                for col in facility_cols
                for value in metadata[col].dropna().tolist()
                if str(value).strip()
            }
        )
        raise ValueError(f"Facility '{facility}' was not found in metadata. Available values: {', '.join(available)}")
    return filtered


def load_client_followups(path):
    if not path or not Path(path).exists():
        return pd.DataFrame(columns=["client", *FOLLOWUP_COLUMNS])
    if str(path).lower().endswith((".xlsx", ".xls")):
        followups = pd.read_excel(path)
    else:
        followups = pd.read_csv(path, sep="\t")
    rename_map = {}
    for col in followups.columns:
        key = normalize(col)
        if key in {"cliente", "client"}:
            rename_map[col] = "client"
        elif key in {"responsable", "owner"}:
            rename_map[col] = "responsable"
        elif key in {"proxima accion", "siguiente accion", "next action"}:
            rename_map[col] = "proxima_accion"
        elif key in {"fecha actualizacion", "fecha de actualizacion", "updated at", "last update"}:
            rename_map[col] = "fecha_actualizacion"
        elif key in {"seguimiento notas", "notas seguimiento", "followup notes"}:
            rename_map[col] = "seguimiento_notas"
    followups = followups.rename(columns=rename_map)
    if "client" not in followups.columns:
        raise ValueError("Client followups must include a 'client' column.")
    followups["client"] = followups["client"].astype(str)
    for col in FOLLOWUP_COLUMNS:
        if col not in followups.columns:
            followups[col] = ""
    return followups[["client", *FOLLOWUP_COLUMNS]]


def filter_followups_by_clients(followups, clients):
    if followups.empty or "client" not in followups:
        return followups
    selected = {str(client) for client in clients}
    return followups[followups["client"].astype(str).isin(selected)].copy()


def enrich_with_client_metadata(df, metadata):
    if df.empty or metadata.empty:
        return df
    metadata = metadata.copy()
    metadata_cols = [col for col in metadata.columns if col != "client"]
    if metadata["client"].duplicated().any():
        rows = []
        for client, group in metadata.groupby(metadata["client"].astype(str), sort=False):
            row = {"client": client}
            facility_values = group[[col for col in FACILITY_COLUMNS if col in group.columns]].fillna("").astype(str)
            has_multiple_facilities = len(facility_values.drop_duplicates()) > 1
            for col in metadata_cols:
                values = [value for value in group[col].tolist() if pd.notna(value) and str(value).strip()]
                if col in FACILITY_COLUMNS and has_multiple_facilities:
                    row[col] = ""
                else:
                    row[col] = values[0] if values else ""
            if has_multiple_facilities:
                notes = [str(value).strip() for value in group.get("notes", pd.Series(dtype=str)).tolist() if pd.notna(value) and str(value).strip()]
                row["notes"] = "Multiples facilities en metadata; usa --facility para aislar covenants. " + " ".join(notes)
            rows.append(row)
        metadata = pd.DataFrame(rows)
    return df.merge(metadata[["client", *metadata_cols]], on="client", how="left")


def discover_documents(clients_root, clients, from_period="", to_period="", max_documents=0, metadata=None, include_undated=False):
    rows = []
    for client in clients:
        financial_dirs = find_financial_dirs(clients_root, client)
        if not financial_dirs:
            base = client_root_dir(clients_root, client) / "1. Data Room/3. Información Financiera/1. Estados Financieros"
            rows.append(
                {
                    "client": client,
                    "path": str(base),
                    "filename": "",
                    "period": "",
                    "statement": "MISSING_DIR",
                    "status": "missing_financial_dir",
                }
            )
            continue
        for base in financial_dirs:
            for path in base.rglob("*"):
                if path.suffix.lower() not in SUPPORTED_SUFFIXES:
                    continue
                statement = classify_statement(path)
                if statement == "UNKNOWN":
                    continue
                period = period_from_name(path)
                if not include_undated and not period:
                    continue
                if from_period and period and period < from_period:
                    continue
                if to_period and period and period > to_period:
                    continue
                rows.append(
                    {
                        "client": client,
                        "path": str(path),
                        "filename": path.name,
                        "file_size_bytes": path.stat().st_size if path.exists() else None,
                        "period": period,
                        "statement": statement,
                        "source_quality": classify_source_quality(path),
                        "statement_frequency": infer_statement_frequency(path),
                        "unit_scale": infer_unit_scale(path),
                        "pair_key": f"{client}|{period}",
                        "status": "indexed" if path.suffix.lower() in EXTRACTABLE_SUFFIXES else "link_only",
                    }
                )
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["client", "period", "statement", "filename"])
        df = df.drop_duplicates(["client", "path"], keep="last")
    if max_documents and len(df) > max_documents:
        df = df.groupby("client", group_keys=False).tail(max(1, max_documents // max(1, len(clients))))
    return enrich_with_client_metadata(df, metadata if metadata is not None else pd.DataFrame())


def append_extra_sources(documents, clients, extra_sources, metadata=None):
    if not extra_sources:
        return documents
    rows = []
    for source in extra_sources:
        for item in str(source).split("|"):
            path = Path(item).expanduser()
            if not item.strip() or not path.exists() or not path.is_file():
                continue
            suffix = path.suffix.lower()
            if suffix not in SUPPORTED_SUFFIXES:
                continue
            matched_clients = [client for client in clients if normalize(client) in normalize(path)]
            client = matched_clients[0] if matched_clients else (clients[0] if len(clients) == 1 else "")
            if not client:
                continue
            period = period_from_name(path)
            statement = classify_statement(path)
            if statement == "UNKNOWN" and suffix in {".xlsx", ".xls"}:
                statement = "EF"
            rows.append(
                {
                    "client": client,
                    "path": str(path),
                    "filename": path.name,
                    "file_size_bytes": path.stat().st_size,
                    "period": period,
                    "statement": statement,
                    "source_quality": classify_source_quality(path),
                    "statement_frequency": infer_statement_frequency(path),
                    "unit_scale": infer_unit_scale(path),
                    "pair_key": f"{client}|{period}",
                    "status": "indexed" if suffix in EXTRACTABLE_SUFFIXES else "link_only",
                }
            )
    if not rows:
        return documents
    extra = enrich_with_client_metadata(pd.DataFrame(rows), metadata if metadata is not None else pd.DataFrame())
    combined = pd.concat([documents, extra], ignore_index=True) if not documents.empty else extra
    return combined.drop_duplicates(["client", "path"], keep="last")


def discover_auxiliary_sources(clients):
    sources = []
    for client in clients:
        client_key = normalize(client)
        matches = []
        for root in AUXILIARY_SOURCE_ROOTS:
            if not root.exists():
                continue
            for path in root.glob("*.xls*"):
                name = normalize(path.name)
                if client_key not in name:
                    continue
                if "modelo" in name or "proyeccion" in name:
                    continue
                if not re.search(r"\beeff\b|estado financiero|estados financieros", name):
                    continue
                matches.append(path)
        if matches:
            sources.append(str(sorted(matches, key=lambda path: path.name, reverse=True)[0]))
    return sources


def pair_label_value_tables(tables):
    rows = []
    index = 0
    while index < len(tables) - 1:
        labels = tables[index]
        values = tables[index + 1]
        labels_are_single = labels and max(len(row) for row in labels if row) == 1
        values_have_amounts = values and max(len(row) for row in values if row) >= 1
        if labels_are_single and values_have_amounts:
            for row_index in range(min(len(labels), len(values))):
                label = (labels[row_index][0] or "").strip()
                amount_cell = values[row_index][0] if values[row_index] else ""
                value = clean_number(amount_cell)
                if label and value is not None:
                    rows.append(
                        {
                            "raw_label": label,
                            "value": value,
                            "raw_value": amount_cell,
                            "source_method": "pdf_table_pair",
                        }
                    )
            index += 2
        else:
            index += 1
    return rows


def _xlsx_number(value):
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)
    if isinstance(value, str):
        return clean_number(value)
    return None


def _xlsx_label(row):
    for value in row[:4]:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _find_period_columns(rows):
    best = []
    for row_idx, row in enumerate(rows[:25], start=1):
        periods = []
        for col_idx, value in enumerate(row, start=1):
            period = period_from_label(value)
            if period:
                periods.append((col_idx, period, value))
        if len(periods) >= 2 and len(periods) > len(best):
            best = periods
    return best


def extract_xlsx_accounts(document, max_rows_per_sheet=400):
    rows = []
    path = Path(document["path"])
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                statement = classify_statement_name(sheet.title)
                if statement == "UNKNOWN":
                    continue
                sheet_rows = []
                for row_no, values in enumerate(sheet.iter_rows(values_only=True), start=1):
                    sheet_rows.append(values)
                    if row_no >= max_rows_per_sheet:
                        break
                period_cols = _find_period_columns(sheet_rows)
                if not period_cols:
                    continue
                for row_no, values in enumerate(sheet_rows, start=1):
                    label = _xlsx_label(values)
                    if not label or normalize(label) in {"concepto", "activo", "pasivo", "capital", "menos"}:
                        continue
                    for col_idx, period, header in period_cols:
                        if col_idx > len(values):
                            continue
                        number = _xlsx_number(values[col_idx - 1])
                        if number is None:
                            continue
                        rows.append(
                            {
                                "raw_label": label,
                                "value": number,
                                "raw_value": f"{sheet.title}!R{row_no}C{col_idx} ({header})",
                                "source_method": "xlsx_period_column",
                                "page": None,
                                "period": period,
                                "statement": statement,
                                "sheet": sheet.title,
                                "row": row_no,
                            }
                        )
        finally:
            workbook.close()
    except Exception as exc:
        return [], str(exc)
    return rows, ""


def extract_pdf_accounts(document, max_pages=5, max_file_mb=8):
    rows = []
    path = Path(document["path"])
    try:
        if path.stat().st_size > max_file_mb * 1024 * 1024:
            return [], f"skipped_heavy_pdf_mb={path.stat().st_size / 1024 / 1024:.1f}"
        with pdfplumber.open(path) as pdf:
            full_text = []
            total_pages = len(pdf.pages)
            if total_pages > max_pages:
                return [], f"skipped_heavy_pdf_pages={total_pages}"
            for page_no, page in enumerate(pdf.pages, start=1):
                use_alarm = _can_use_signal_alarm()
                if use_alarm:
                    signal.signal(signal.SIGALRM, timeout_handler)
                    signal.alarm(PAGE_TIMEOUT_SECONDS)
                try:
                    page_text = page.extract_text() or ""
                except PageTimeout as exc:
                    rows.append(
                        {
                            "raw_label": "ERROR",
                            "value": None,
                            "raw_value": str(exc),
                            "source_method": "extract_error",
                            "page": page_no,
                        }
                    )
                    continue
                finally:
                    if use_alarm:
                        signal.alarm(0)
                full_text.append(page_text)
                if use_alarm:
                    signal.alarm(PAGE_TIMEOUT_SECONDS)
                try:
                    tables = page.extract_tables() or []
                except PageTimeout as exc:
                    rows.append(
                        {
                            "raw_label": "ERROR",
                            "value": None,
                            "raw_value": str(exc),
                            "source_method": "extract_error",
                            "page": page_no,
                        }
                    )
                    tables = []
                except Exception:
                    tables = []
                finally:
                    if use_alarm:
                        signal.alarm(0)
                for account in pair_label_value_tables(tables):
                    account.update({"page": page_no})
                    rows.append(account)
            text = "\n".join(full_text)
    except Exception as exc:
        return [], str(exc)

    for label in REQUIRED_TOTAL_LABELS:
        for line in text.splitlines():
            if label.lower() in line.lower():
                value = first_money_after_label(line, label)
                if value is not None:
                    rows.append(
                        {
                            "raw_label": label,
                            "value": value,
                            "raw_value": line,
                            "source_method": "pdf_text_total",
                            "page": None,
                        }
                    )
                break
    return rows, ""


def extract_pdf_accounts_cached(document, max_pages=5, max_file_mb=8, refresh_cache=False):
    path = Path(document["path"])
    if not refresh_cache:
        cached = read_pdf_cache(path, max_pages, max_file_mb)
        if cached is not None:
            return cached.get("rows", []), cached.get("error", ""), True
    rows, error = extract_pdf_accounts(document, max_pages=max_pages, max_file_mb=max_file_mb)
    try:
        write_pdf_cache(path, max_pages, max_file_mb, rows, error)
    except OSError:
        pass
    return rows, error, False


def _extract_pdf_accounts_task(document, max_pages, max_file_mb, refresh_cache):
    rows, error, cache_hit = extract_pdf_accounts_cached(
        document,
        max_pages=max_pages,
        max_file_mb=max_file_mb,
        refresh_cache=refresh_cache,
    )
    return rows, error, cache_hit


def _run_pdf_tasks_in_executor(executor_class, pending, worker_count, max_pages, max_file_mb, refresh_cache):
    results = {}
    with executor_class(max_workers=worker_count) as executor:
        futures = {
            executor.submit(_extract_pdf_accounts_task, document, max_pages, max_file_mb, refresh_cache): (
                position,
                document,
            )
            for position, document in pending
        }
        for future in as_completed(futures):
            position, document = futures[future]
            try:
                rows, error, cache_hit = future.result()
            except Exception as exc:
                rows, error, cache_hit = [], str(exc), False
            results[position] = (document, rows, error, cache_hit)
    return results


def _account_rows_for_document(document, rows, error):
    if error:
        return [
            {
                **document,
                "raw_label": "",
                "normalized_label": "",
                "value": None,
                "raw_value": "",
                "page": None,
                "source_method": "error",
                "extract_error": error,
            }
        ]
    account_rows = []
    for row in rows:
        unit_scale = document.get("unit_scale", 1) or 1
        account_document = {
            **document,
            "period": row.get("period") or document.get("period"),
            "statement": row.get("statement") or document.get("statement"),
        }
        source_suffix = f"#{row.get('sheet')}!R{row.get('row')}" if row.get("sheet") and row.get("row") else ""
        account_rows.append(
            {
                **account_document,
                "raw_label": row["raw_label"],
                "normalized_label": normalize(row["raw_label"]),
                "value": row["value"] * unit_scale if row["value"] is not None else None,
                "original_value": row["value"],
                "unit_scale_applied": unit_scale,
                "raw_value": row["raw_value"],
                "page": row["page"],
                "source_method": row["source_method"],
                "source_ref": f"{document.get('path')}#page={row['page']}" if row.get("page") else f"{document.get('path')}{source_suffix}",
                "extract_error": "",
            }
        )
    return account_rows


def extract_accounts(documents, max_pages=5, max_file_mb=8, refresh_cache=False, pdf_workers=1):
    account_rows = []
    cache_hits = 0
    pdf_documents = []
    xlsx_count = 0
    for position, document in documents.iterrows():
        document = document.to_dict()
        suffix = Path(document["path"]).suffix.lower()
        if suffix == ".pdf":
            pdf_documents.append((position, document))
        elif suffix in {".xlsx", ".xls"}:
            xlsx_count += 1
            rows, error = extract_xlsx_accounts(document)
            account_rows.extend(_account_rows_for_document(document, rows, error))

    pdf_count = len(pdf_documents)
    if pdf_count == 0:
        accounts = pd.DataFrame(account_rows)
        accounts.attrs["pdf_count"] = pdf_count
        accounts.attrs["cache_hits"] = cache_hits
        accounts.attrs["xlsx_count"] = xlsx_count
        return accounts

    pdf_workers = max(1, int(pdf_workers or 1))
    if pdf_workers == 1 or pdf_count == 1:
        results = []
        for position, document in pdf_documents:
            rows, error, cache_hit = extract_pdf_accounts_cached(
                document,
                max_pages=max_pages,
                max_file_mb=max_file_mb,
                refresh_cache=refresh_cache,
            )
            results.append((position, document, rows, error, cache_hit))
    else:
        results_by_position = {}
        pending = []
        if refresh_cache:
            pending = pdf_documents
        else:
            for position, document in pdf_documents:
                cached = read_pdf_cache(Path(document["path"]), max_pages, max_file_mb)
                if cached is None:
                    pending.append((position, document))
                else:
                    results_by_position[position] = (
                        document,
                        cached.get("rows", []),
                        cached.get("error", ""),
                        True,
                    )

        worker_count = min(pdf_workers, len(pending))
        if worker_count:
            try:
                results_by_position.update(
                    _run_pdf_tasks_in_executor(
                        ProcessPoolExecutor,
                        pending,
                        worker_count,
                        max_pages,
                        max_file_mb,
                        refresh_cache,
                    )
                )
            except (OSError, PermissionError):
                results_by_position.update(
                    _run_pdf_tasks_in_executor(
                        ThreadPoolExecutor,
                        pending,
                        worker_count,
                        max_pages,
                        max_file_mb,
                        refresh_cache,
                    )
                )

        results = [
            (position, *results_by_position[position])
            for position, _document in sorted(pdf_documents, key=lambda item: item[0])
        ]

    for _position, document, rows, error, cache_hit in results:
        cache_hits += int(cache_hit)
        account_rows.extend(_account_rows_for_document(document, rows, error))
    accounts = pd.DataFrame(account_rows)
    accounts.attrs["pdf_count"] = pdf_count
    accounts.attrs["cache_hits"] = cache_hits
    accounts.attrs["xlsx_count"] = xlsx_count
    return accounts


def load_mapping_memory():
    mapping = pd.read_csv(MAPPING_PATH, sep="\t")
    mapping["normalized_label"] = mapping["raw_label"].map(normalize)
    return mapping


def build_mapping_index(mapping):
    priority = {"approved": 0, "proxy": 1}
    ordered = mapping.copy()
    ordered["_status_rank"] = ordered["status"].map(priority).fillna(9)
    ordered = ordered.sort_values(["_status_rank", "confidence"], ascending=[True, False])
    index = {}
    for row in ordered.to_dict("records"):
        label = row["normalized_label"]
        keys = [(label, row["client"]), (label, "*")]
        if row["entity_type"] == "generic":
            keys.append((label, "__generic__"))
        for key in keys:
            index.setdefault(key, row)
    return index


def map_accounts(accounts, mapping):
    if accounts.empty:
        mapped_df = accounts.copy()
        for col, default in {
            "concept": "",
            "mapping_confidence": 0.0,
            "mapping_status": "unmapped",
            "mapping_notes": "",
        }.items():
            mapped_df[col] = default
        return mapped_df
    mapping_index = build_mapping_index(mapping)
    mapped = []
    for _, account in accounts.iterrows():
        label = account["normalized_label"]
        candidate = (
            mapping_index.get((label, account["client"]))
            or mapping_index.get((label, "*"))
            or mapping_index.get((label, "__generic__"))
        )
        if candidate is None:
            concept = ""
            confidence = 0.0
            mapping_status = "unmapped"
            mapping_notes = ""
        else:
            concept = candidate["concept"]
            confidence = candidate["confidence"]
            mapping_status = candidate["status"]
            mapping_notes = "" if pd.isna(candidate.get("notes")) else candidate.get("notes")
        mapped.append(
            {
                **account.to_dict(),
                "concept": concept,
                "mapping_confidence": confidence,
                "mapping_status": mapping_status,
                "mapping_notes": mapping_notes,
            }
        )
    mapped_df = pd.DataFrame(mapped)
    mapped_df.attrs.update(accounts.attrs)
    return mapped_df


def best_concepts(mapped):
    usable = mapped[mapped["concept"].fillna("").ne("")].copy()
    if usable.empty:
        return pd.DataFrame()
    usable["rank"] = usable["mapping_confidence"].fillna(0)
    usable = usable.sort_values(["client", "period", "concept", "rank"], ascending=[True, True, True, False])
    return usable.drop_duplicates(["client", "period", "concept"], keep="first")


def value(concepts, client, period, concept):
    subset = concepts[
        concepts["client"].eq(client) & concepts["period"].eq(period) & concepts["concept"].eq(concept)
    ]
    if subset.empty:
        return None
    return subset.iloc[0]["value"]


def concept_source(concepts, client, period, concept):
    subset = concepts[
        concepts["client"].eq(client) & concepts["period"].eq(period) & concepts["concept"].eq(concept)
    ]
    if subset.empty:
        return {}
    row = subset.iloc[0]
    return {
        "value": row.get("value"),
        "raw_label": row.get("raw_label"),
        "filename": row.get("filename"),
        "source_ref": row.get("source_ref"),
        "mapping_confidence": row.get("mapping_confidence"),
        "mapping_status": row.get("mapping_status"),
    }


def allowance_adjustment(value):
    if value is None or pd.isna(value):
        return 0
    return -abs(value)


def calc_ratio(numerator, denominator):
    if numerator is None or denominator in (None, 0):
        return None
    return numerator / denominator


def abs_value(value):
    if value is None or pd.isna(value):
        return None
    return abs(value)


def client_meta_for(concepts, client):
    subset = concepts[concepts["client"].eq(client)]
    if subset.empty:
        return {}
    row = subset.iloc[0]
    return {col: row[col] for col in subset.columns if col in {
        "tipo_entidad_juridica",
        "antiguedad_operativa",
        "producto_principal",
        "segmento_objetivo",
        "modelo_fondeo",
        "cobertura_geografica",
        "canal_originacion",
        "etapa_madurez",
        "ticket_promedio_rango",
        "tipo_estado_financiero",
        "anios_informacion_financiera",
        "se_otorgo_credito",
        "contrato_drive_path",
        "contrato_drive_link",
        "facility_id",
        "facility_name",
        "facility_covenants",
    }}


def ratio_policy(ratio_name, meta):
    product = normalize(meta.get("producto_principal", ""))
    entity = normalize(meta.get("tipo_entidad_juridica", ""))
    notes = []
    status_override = None
    if ratio_name in {"Cobertura de Deuda", "ICAP Ajustado"}:
        status_override = "needs_review"
        notes.append("Definicion depende de tipo de entidad/producto.")
    if ratio_name in {"Tasa Activa", "Margen Financiero"} and "arrendamiento" in product:
        status_override = "needs_review"
        notes.append("Producto arrendamiento: validar si ingresos por rentas se tratan como intereses/ingreso financiero.")
    if ratio_name == "ICAP" and entity in {"sofom", "sofipo"}:
        notes.append("Validar si ICAP regulatorio requiere metodologia distinta a capital contable / activo.")
        status_override = "needs_review"
    return status_override, " ".join(notes)


def pairing_qa(documents):
    if documents.empty:
        return pd.DataFrame()
    valid = documents[documents["status"].eq("indexed") & documents["period"].fillna("").ne("")]
    rows = []
    for (client, period), group in valid.groupby(["client", "period"]):
        statements = sorted(set(group["statement"].dropna()))
        rows.append(
            {
                "client": client,
                "period": period,
                "check": "statement_pairing",
                "difference": "",
                "status": "ok" if {"BG", "ER"}.issubset(set(statements)) or "EF" in statements else "needs_review",
                "details": ", ".join(statements),
            }
        )
    return pd.DataFrame(rows)


def calculate_ratios(concepts):
    rows = []
    qa_rows = []
    for client in sorted(concepts["client"].dropna().unique()):
        meta = client_meta_for(concepts, client)
        for period in sorted(concepts[concepts["client"].eq(client)]["period"].dropna().unique()):
            total_activo = value(concepts, client, period, "total_activo")
            total_pasivo = value(concepts, client, period, "total_pasivo")
            total_capital = value(concepts, client, period, "total_capital_contable")
            balance_diff = None
            if total_activo is not None and total_pasivo is not None and total_capital is not None:
                balance_diff = total_activo - total_pasivo - total_capital
            qa_status = "ok" if balance_diff is not None and abs(balance_diff) <= 1 else "needs_review"
            qa_rows.append(
                {
                    "client": client,
                    "period": period,
                    "check": "total_activo = total_pasivo + total_capital_contable",
                    "difference": balance_diff,
                    "status": qa_status,
                    "details": "",
                }
            )

            cartera_neta = None
            arr = value(concepts, client, period, "clientes_arrendamiento")
            fac = value(concepts, client, period, "clientes_factoraje")
            est = value(concepts, client, period, "estimacion_preventiva")
            if arr is not None or fac is not None:
                cartera_neta = (arr or 0) + (fac or 0) + allowance_adjustment(est)

            deuda = (value(concepts, client, period, "fondeadores_cp") or 0) + (
                value(concepts, client, period, "fondeadores_lp") or 0
            )
            if deuda == 0:
                deuda = (value(concepts, client, period, "deuda_cp_proxy") or 0) + (
                    value(concepts, client, period, "deuda_lp_proxy") or 0
                )
                deuda_note = "Deuda usa proxy; validar."
            else:
                deuda_note = ""

            activos_productivos = None
            efectivo = value(concepts, client, period, "efectivo_inversiones")
            otros = concepts[
                concepts["client"].eq(client)
                & concepts["period"].eq(period)
                & concepts["concept"].eq("otros_activos_generadores")
            ]["value"].sum()
            dep = concepts[
                concepts["client"].eq(client)
                & concepts["period"].eq(period)
                & concepts["concept"].eq("depreciacion_activos_productivos")
            ]["value"].sum()
            if efectivo is not None or cartera_neta is not None or otros:
                activos_productivos = (efectivo or 0) + (cartera_neta or 0) + otros + dep

            ingresos_intereses = value(concepts, client, period, "ingresos_intereses")
            ingresos_comisiones = concepts[
                concepts["client"].eq(client)
                & concepts["period"].eq(period)
                & concepts["concept"].eq("ingresos_por_comisiones")
            ]["value"].sum()
            gastos_operacion = value(concepts, client, period, "gastos_operacion")
            cartera_vigente = value(concepts, client, period, "cartera_vigente")
            cartera_atrasada = value(concepts, client, period, "cartera_atrasada")
            cartera_vencida = value(concepts, client, period, "cartera_vencida")
            cartera_bruta = value(concepts, client, period, "cartera_bruta")
            cartera_neta_credito = value(concepts, client, period, "cartera_neta_credito")
            estimacion_preventiva = abs_value(value(concepts, client, period, "estimacion_preventiva"))
            cartera_total_contractual = None
            if cartera_bruta is not None:
                cartera_total_contractual = cartera_bruta
            elif cartera_vigente is not None or cartera_atrasada is not None or cartera_vencida is not None:
                cartera_total_contractual = (cartera_vigente or 0) + (cartera_atrasada or 0) + (cartera_vencida or 0)
            ingresos_operativos = None
            if ingresos_intereses is not None or ingresos_comisiones:
                ingresos_operativos = (ingresos_intereses or 0) + ingresos_comisiones

            specs = [
                ("Margen Operativo", value(concepts, client, period, "utilidad_operacion"), value(concepts, client, period, "ingresos_totales"), "utilidad_operacion / ingresos_totales", "calculated", ""),
                ("Margen Neto", value(concepts, client, period, "utilidad_neta"), value(concepts, client, period, "ingresos_totales"), "utilidad_neta / ingresos_totales", "calculated", ""),
                ("ROE", value(concepts, client, period, "utilidad_neta"), total_capital, "utilidad_neta / total_capital_contable", "calculated", ""),
                ("ROA", value(concepts, client, period, "utilidad_neta"), total_activo, "utilidad_neta / total_activo", "calculated", ""),
                ("Tasa Pasiva", value(concepts, client, period, "gastos_financieros"), deuda, "gastos_financieros / deuda", "calculated", deuda_note),
                ("Apalancamiento", deuda, total_activo, "deuda / total_activo", "calculated", deuda_note),
                ("ICAP", total_capital, total_activo, "total_capital_contable / total_activo", "calculated", ""),
                ("ICAP Ajustado", total_capital, cartera_neta, "total_capital_contable / cartera_neta", "needs_review", "Cartera neta derivada; validar definicion."),
                ("Cobertura de Deuda", activos_productivos, total_pasivo, "activos_productivos / total_pasivo", "needs_review", "Activos productivos derivado; validar definicion."),
                ("Rentabilidad Operativa", abs_value(gastos_operacion), ingresos_operativos, "abs(gastos_operacion) / (ingresos_intereses + ingresos_por_comisiones)", "calculated", "Covenant Cofine segun ficha contractual."),
                ("Cartera Vencida Neta", (cartera_vencida or 0) - (estimacion_preventiva or 0) if cartera_vencida is not None or estimacion_preventiva is not None else None, cartera_total_contractual, "(cartera_vencida - estimacion_preventiva) / cartera_total", "calculated", "Covenant Cofine segun ficha contractual."),
                ("Capitalización", cartera_neta_credito, total_capital, "cartera_neta_credito / total_capital_contable", "calculated", "Covenant Cofine segun ficha contractual."),
                ("Estimación preventiva de riesgos", estimacion_preventiva, cartera_vencida, "estimacion_preventiva / cartera_vencida", "calculated", "Covenant Cofine segun ficha contractual."),
            ]
            for name, numerator, denominator, formula, base_status, note in specs:
                result = calc_ratio(numerator, denominator)
                missing = []
                if numerator is None:
                    missing.append("numerador")
                if denominator in (None, 0):
                    missing.append("denominador")
                status = base_status
                notes = [note] if note else []
                policy_status, policy_note = ratio_policy(name, meta)
                if policy_status:
                    status = policy_status
                if policy_note:
                    notes.append(policy_note)
                if qa_status != "ok":
                    status = "needs_review"
                    notes.append("QA balance no ok.")
                if missing:
                    status = "needs_review"
                    notes.append("Inputs faltantes: " + ", ".join(missing))
                if result is not None and abs(result) > 10:
                    status = "needs_review"
                    notes.append("Resultado > 10x; validar.")
                rows.append(
                    {
                        "client": client,
                        "period": period,
                        "ratio": name,
                        "numerator": numerator,
                        "denominator": denominator,
                        "result": result,
                        "result_pct": None if result is None else result * 100,
                        "formula": formula,
                        "review_status": status,
                        "review_notes": " ".join(notes),
                        **meta,
                    }
                )
    return pd.DataFrame(rows), pd.DataFrame(qa_rows)


def _selected_covenant_names(value):
    if value is None or pd.isna(value) or not str(value).strip():
        return set()
    return {normalize(part) for part in re.split(r"[|,;/]+", str(value)) if normalize(part)}


def filter_ratios_by_facility_covenants(ratios, metadata, facility=""):
    if ratios.empty or metadata.empty or "facility_covenants" not in metadata.columns:
        return ratios
    covenant_map = {}
    for client, group in metadata.groupby(metadata["client"].astype(str), sort=False):
        covenant_sets = [
            _selected_covenant_names(row.get("facility_covenants"))
            for _, row in group.iterrows()
            if _selected_covenant_names(row.get("facility_covenants"))
        ]
        unique_sets = {tuple(sorted(items)) for items in covenant_sets}
        if not unique_sets:
            continue
        if facility or len(unique_sets) == 1:
            covenant_map[str(client)] = set().union(*covenant_sets)
    if not covenant_map:
        return ratios

    def keep(row):
        selected = covenant_map.get(str(row["client"]))
        if not selected:
            return True
        return normalize(row.get("ratio")) in selected

    return ratios[ratios.apply(keep, axis=1)].copy()


def facility_selection_qa(metadata, facility=""):
    if metadata.empty or "client" not in metadata or "facility_covenants" not in metadata:
        return pd.DataFrame()
    rows = []
    for client, group in metadata.groupby(metadata["client"].astype(str), sort=False):
        covenant_sets = [
            tuple(sorted(_selected_covenant_names(row.get("facility_covenants"))))
            for _, row in group.iterrows()
            if _selected_covenant_names(row.get("facility_covenants"))
        ]
        if not covenant_sets:
            rows.append(
                {
                    "client": client,
                    "period": "",
                    "check": "facility_covenants_configured",
                    "difference": "",
                    "status": "needs_review",
                    "details": "No hay facility_covenants configurados; no se puede elegir covenants de facility.",
                }
            )
        elif not facility and len(set(covenant_sets)) > 1:
            rows.append(
                {
                    "client": client,
                    "period": "",
                    "check": "facility_filter_required",
                    "difference": "",
                    "status": "needs_review",
                    "details": "Hay varias facilities con covenants distintos; corre con --facility para no mezclar.",
                }
            )
    return pd.DataFrame(rows)


def golden_sample_scaffold(ratios):
    if ratios.empty:
        return pd.DataFrame()
    sample = ratios[["client", "period", "ratio", "result", "formula", "review_status"]].copy()
    sample = sample.sort_values(["client", "period", "ratio"]).groupby(["client", "period"]).head(5)
    sample["manual_expected_result"] = ""
    sample["manual_expected_pct"] = ""
    sample["reviewer"] = ""
    sample["approval_status"] = "pending"
    sample["review_notes"] = ""
    return sample


def build_loan_tape_by_client(concepts):
    columns = [
        "Cliente",
        "Periodo",
        "Base analisis",
        "Cartera arrendamiento",
        "Cartera factoraje",
        "Cartera bruta",
        "Estimacion preventiva reportada",
        "Ajuste estimacion usado",
        "Cartera neta analizada",
        "Estimacion / cartera bruta",
        "Calidad cartera",
        "Alertas cartera",
        "Fuente arrendamiento",
        "Fuente factoraje",
        "Fuente estimacion",
        "Formula",
    ]
    if concepts.empty or "client" not in concepts or "period" not in concepts:
        return pd.DataFrame(columns=columns)

    rows = []
    for client in sorted(concepts["client"].dropna().astype(str).unique()):
        periods = sorted(concepts[concepts["client"].astype(str).eq(client)]["period"].dropna().astype(str).unique())
        for period in periods:
            arr = concept_source(concepts, client, period, "clientes_arrendamiento")
            fac = concept_source(concepts, client, period, "clientes_factoraje")
            est = concept_source(concepts, client, period, "estimacion_preventiva")
            arr_value = arr.get("value")
            fac_value = fac.get("value")
            est_value = est.get("value")
            gross = None
            if arr_value is not None or fac_value is not None:
                gross = (arr_value or 0) + (fac_value or 0)
            adjustment = allowance_adjustment(est_value)
            net = None if gross is None else gross + adjustment
            allowance_ratio = None if not gross else abs(adjustment) / gross

            alerts = ["Proxy contable: no sustituye loan tape granular."]
            quality = "OK proxy"
            if gross is None:
                quality = "Sin cartera"
                alerts.append("No se detectaron cuentas de cartera para el periodo.")
            elif not arr and not fac:
                quality = "Revisar"
                alerts.append("No hay desglose arrendamiento/factoraje.")
            elif not est:
                quality = "Revisar"
                alerts.append("Falta estimacion preventiva; cartera neta usa cartera bruta.")
            if net is not None and net < 0:
                quality = "Revisar"
                alerts.append("Cartera neta negativa; validar extraccion.")

            def source_label(source):
                if not source:
                    return ""
                source_ref = str(source.get("source_ref") or "").strip()
                page_ref = ""
                if "#page=" in source_ref:
                    page_ref = "#" + source_ref.split("#", 1)[1]
                filename = str(source.get("filename") or "").strip()
                if not filename and source_ref:
                    filename = Path(source_ref.split("#", 1)[0]).name
                parts = [
                    str(source.get("raw_label") or "").strip(),
                    f"{filename}{page_ref}" if filename else page_ref,
                ]
                return " | ".join(part for part in parts if part)

            rows.append(
                {
                    "Cliente": client,
                    "Periodo": period,
                    "Base analisis": "Proxy contable desde BG/EEFF",
                    "Cartera arrendamiento": arr_value,
                    "Cartera factoraje": fac_value,
                    "Cartera bruta": gross,
                    "Estimacion preventiva reportada": est_value,
                    "Ajuste estimacion usado": adjustment if est_value is not None else None,
                    "Cartera neta analizada": net,
                    "Estimacion / cartera bruta": allowance_ratio,
                    "Calidad cartera": quality,
                    "Alertas cartera": " ".join(alerts),
                    "Fuente arrendamiento": source_label(arr),
                    "Fuente factoraje": source_label(fac),
                    "Fuente estimacion": source_label(est),
                    "Formula": "arrendamiento + factoraje - abs(estimacion_preventiva)",
                }
            )
    return pd.DataFrame(rows, columns=columns)


def _source_label_from_concept(row):
    if row is None or row.empty:
        return ""
    source_ref = str(row.get("source_ref") or "").strip()
    page_ref = ""
    if "#page=" in source_ref:
        page_ref = "#" + source_ref.split("#", 1)[1]
    filename = str(row.get("filename") or "").strip()
    if not filename and source_ref:
        filename = Path(source_ref.split("#", 1)[0]).name
    parts = [
        str(row.get("raw_label") or "").strip(),
        f"{filename}{page_ref}" if filename else page_ref,
    ]
    return " | ".join(part for part in parts if part)


def build_financial_statements_view(concepts):
    base_columns = ["Cliente", "Estado", "Categoria", "Concepto", "Concepto normalizado"]
    audit_columns = ["Fuente ultimo periodo", "Etiqueta fuente", "Mapping", "Notas"]
    if concepts.empty or "client" not in concepts or "period" not in concepts:
        return pd.DataFrame(columns=[*base_columns, *audit_columns])

    concepts = concepts.copy()
    concepts["client"] = concepts["client"].astype(str)
    concepts["period"] = concepts["period"].astype(str)
    periods = sorted(period for period in concepts["period"].dropna().unique() if period)
    rows = []

    for client in sorted(concepts["client"].dropna().unique()):
        client_concepts = concepts[concepts["client"].eq(client)]
        for statement, category, label, concept in FINANCIAL_STATEMENT_LAYOUT:
            subset = client_concepts[client_concepts["concept"].eq(concept)].sort_values("period")
            row = {
                "Cliente": client,
                "Estado": statement,
                "Categoria": category,
                "Concepto": label,
                "Concepto normalizado": concept,
            }
            for period in periods:
                period_subset = subset[subset["period"].eq(period)]
                row[period] = None if period_subset.empty else period_subset.iloc[0].get("value")
            if subset.empty:
                latest = None
                notes = "No mapeado en periodos procesados."
            else:
                latest = subset.iloc[-1]
                notes = latest.get("mapping_notes")
            row.update(
                {
                    "Fuente ultimo periodo": "" if latest is None else latest.get("filename", ""),
                    "Etiqueta fuente": _source_label_from_concept(latest),
                    "Mapping": "" if latest is None else latest.get("mapping_status", ""),
                    "Notas": "" if pd.isna(notes) else notes,
                }
            )
            rows.append(row)

    return pd.DataFrame(rows, columns=[*base_columns, *periods, *audit_columns])


def _first_nonblank(series, default=""):
    if series is None:
        return default
    cleaned = [value for value in series.tolist() if pd.notna(value) and str(value).strip()]
    return cleaned[0] if cleaned else default


def _client_groups(frame):
    if frame.empty or "client" not in frame:
        return {}
    return {str(client): group for client, group in frame.groupby(frame["client"].astype(str), sort=False)}


def _latest_loan_tape_by_client(loan_tape):
    if loan_tape is None or loan_tape.empty or "Cliente" not in loan_tape:
        return {}
    latest = {}
    for client, group in loan_tape.groupby(loan_tape["Cliente"].astype(str), sort=False):
        group = group.sort_values("Periodo")
        latest[str(client)] = group.iloc[-1].to_dict()
    return latest


def build_crm_clients(documents, ratios, qa, followups=None, loan_tape=None):
    clients = set()
    if not documents.empty and "client" in documents:
        clients.update(documents["client"].dropna().astype(str))
    if not ratios.empty and "client" in ratios:
        clients.update(ratios["client"].dropna().astype(str))
    if not qa.empty and "client" in qa:
        clients.update(qa["client"].dropna().astype(str))
    if followups is not None and not followups.empty and "client" in followups:
        clients.update(followups["client"].dropna().astype(str))

    document_groups = _client_groups(documents)
    ratio_groups = _client_groups(ratios)
    qa_groups = _client_groups(qa)
    followup_groups = _client_groups(followups if followups is not None else pd.DataFrame())
    loan_tape_latest = _latest_loan_tape_by_client(loan_tape)
    rows = []
    for client in sorted(clients):
        doc_client = document_groups.get(client, pd.DataFrame())
        ratio_client = ratio_groups.get(client, pd.DataFrame())
        qa_client = qa_groups.get(client, pd.DataFrame())
        followup_client = followup_groups.get(client, pd.DataFrame())
        loan_client = loan_tape_latest.get(client, {})

        periods = sorted(doc_client["period"].dropna().astype(str).unique()) if "period" in doc_client else []
        latest_period = periods[-1] if periods else ""
        ratio_review = int(ratio_client["review_status"].eq("needs_review").sum()) if "review_status" in ratio_client else 0
        qa_review = int(qa_client["status"].eq("needs_review").sum()) if "status" in qa_client else 0
        documents_found = len(doc_client)
        status = "Listo para revisar"
        priority = "Media"
        blocker = "Listo"
        suggested_action = "Validar y actualizar proxima accion"
        if documents_found == 0:
            status = "Sin documentos"
            priority = "Alta"
            blocker = "Faltan estados financieros"
            suggested_action = "Agregar EEFF y reprocesar"
        elif ratio_review or qa_review:
            status = "Requiere revision"
            priority = "Alta"
            blocker_parts = []
            if ratio_review:
                blocker_parts.append(f"{ratio_review} razones")
            if qa_review:
                blocker_parts.append(f"{qa_review} checks QA")
            blocker = "Revisar " + " + ".join(blocker_parts)
            suggested_action = "Abrir hojas Razones/QA y resolver alertas"
        elif ratio_client.empty:
            status = "Sin razones"
            priority = "Media"
            blocker = "Sin razones calculadas"
            suggested_action = "Validar mapeo de cuentas"
        else:
            status = "Actualizado"
            priority = "Baja"
            suggested_action = "Confirmar seguimiento comercial"

        rows.append(
            {
                "Cliente": client,
                "Estatus": status,
                "Prioridad": priority,
                "Bloqueo principal": blocker,
                "Accion sugerida": suggested_action,
                "Responsable": _first_nonblank(followup_client.get("responsable")),
                "Proxima accion": _first_nonblank(followup_client.get("proxima_accion")),
                "Fecha actualizacion": _first_nonblank(followup_client.get("fecha_actualizacion")),
                "Ultimo periodo": latest_period,
                "Facility ID": _first_nonblank(doc_client.get("facility_id")) or _first_nonblank(ratio_client.get("facility_id")),
                "Facility": _first_nonblank(doc_client.get("facility_name")) or _first_nonblank(ratio_client.get("facility_name")),
                "Covenants facility": _first_nonblank(doc_client.get("facility_covenants")) or _first_nonblank(ratio_client.get("facility_covenants")),
                "Docs": documents_found,
                "Cartera neta": loan_client.get("Cartera neta analizada", ""),
                "Calidad cartera": loan_client.get("Calidad cartera", ""),
                "Base cartera": loan_client.get("Base analisis", ""),
                "Alertas cartera": loan_client.get("Alertas cartera", ""),
                "Razones revisar": ratio_review,
                "QA revisar": qa_review,
                "Credito": _first_nonblank(doc_client.get("se_otorgo_credito")),
                "Producto": _first_nonblank(doc_client.get("producto_principal")),
                "Link contrato": _first_nonblank(doc_client.get("contrato_drive_link")),
                "Notas seguimiento": _first_nonblank(followup_client.get("seguimiento_notas")),
                "Notas": _first_nonblank(doc_client.get("notes")),
            }
        )
    return pd.DataFrame(rows)


def build_inicio(crm, documents, accounts, concepts, ratios, qa, loan_tape=None):
    if crm.empty:
        return pd.DataFrame(
            [
                {"Seccion": "Estado", "Metrica": "Clientes", "Valor": 0, "Detalle": "No se encontraron clientes en la corrida."},
                {"Seccion": "Siguiente paso", "Metrica": "Carga", "Valor": "", "Detalle": "Revisar ruta de clientes y volver a ejecutar."},
            ]
        )
    return pd.DataFrame(
        [
            {"Seccion": "Estado", "Metrica": "Clientes", "Valor": len(crm), "Detalle": "Filas en CRM Clientes."},
            {"Seccion": "Estado", "Metrica": "Actualizados", "Valor": int(crm["Estatus"].eq("Actualizado").sum()), "Detalle": "Sin alertas de razones ni QA."},
            {"Seccion": "Estado", "Metrica": "Requieren revision", "Valor": int(crm["Estatus"].eq("Requiere revision").sum()), "Detalle": "Tienen razones o QA por validar."},
            {"Seccion": "Estado", "Metrica": "Sin documentos", "Valor": int(crm["Estatus"].eq("Sin documentos").sum()), "Detalle": "No se indexaron estados financieros."},
            {"Seccion": "Pipeline", "Metrica": "Documentos", "Valor": len(documents), "Detalle": "Archivos financieros indexados."},
            {"Seccion": "Pipeline", "Metrica": "Cuentas extraidas", "Valor": len(accounts), "Detalle": "Renglones leidos desde PDFs/Excel."},
            {"Seccion": "Pipeline", "Metrica": "Conceptos mapeados", "Valor": len(concepts), "Detalle": "Cuentas normalizadas para calculo."},
            {"Seccion": "Pipeline", "Metrica": "Razones", "Valor": len(ratios), "Detalle": "Razones financieras calculadas."},
            {
                "Seccion": "Pipeline",
                "Metrica": "QA pendientes",
                "Valor": int(qa["status"].eq("needs_review").sum()) if not qa.empty and "status" in qa else 0,
                "Detalle": "Checks que requieren revision.",
            },
            {
                "Seccion": "Loan tape",
                "Metrica": "Periodos analizados",
                "Valor": len(loan_tape) if loan_tape is not None else 0,
                "Detalle": "Proxy contable por cliente; validar contra loan tape granular si existe.",
            },
        ]
    )


def _json_ready(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):
        value = value.item()
    if hasattr(value, "isoformat"):
        return value.isoformat()[:10]
    return value


def _records_for_json(frame, max_rows=None):
    if frame is None or frame.empty:
        return []
    records = [
        {str(key): _json_ready(value) for key, value in row.items()}
        for row in frame.to_dict("records")
    ]
    return records[:max_rows] if max_rows else records


def write_crm_sidecar(
    path,
    crm,
    ratios=None,
    qa=None,
    documents=None,
    concepts=None,
    loan_tape=None,
    financial_statements=None,
):
    sidecar_path = path.with_suffix(".crm.json")
    rows = _records_for_json(crm)
    payload = {
        "source_workbook": str(path),
        "source_mtime_ns": path.stat().st_mtime_ns if path.exists() else None,
        "rows": rows,
        "ratios": _records_for_json(ratios),
        "qa": _records_for_json(qa),
        "documents": _records_for_json(documents),
        "concepts": _records_for_json(concepts, max_rows=5000),
        "loanTape": _records_for_json(loan_tape),
        "financialStatements": _records_for_json(financial_statements),
    }
    sidecar_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def build_update_guide(output_path):
    rows = [
        {"Paso": 1, "Accion": "Dashboard Comercial", "Detalle": "Revisa Inicio para KPIs, pipeline, cartera y semaforos ejecutivos."},
        {"Paso": 2, "Accion": "CRM Comercial", "Detalle": "Filtra CRM Clientes por Prioridad Alta, asigna Responsable y actualiza Proxima accion / Fecha actualizacion."},
        {"Paso": 3, "Accion": "Monitoreo", "Detalle": "Usa Linea de vida / Clientes para revisar alertas, bloqueo principal y cambios relevantes por cliente."},
        {"Paso": 4, "Accion": "Linea de Analisis", "Detalle": "Revisa benchmarking, consolidacion y Z Core / Z-Score desde Razones y QA; filtra needs_review."},
        {"Paso": 5, "Accion": "MI Quality", "Detalle": "Controla ingestion, validacion y consolidacion en Actualizar, Documentos y Auditoria."},
        {"Paso": 6, "Accion": "Editar clientes", "Detalle": "Actualiza config/client_metadata_template.tsv para campos permanentes como producto, credito, contrato y notas."},
        {"Paso": 7, "Accion": "Editar seguimiento", "Detalle": "Actualiza config/client_followups.tsv para conservar Responsable / Proxima accion / Fecha actualizacion / Notas seguimiento entre corridas."},
        {"Paso": 8, "Accion": "Agregar documentos", "Detalle": "Coloca estados financieros en la carpeta del cliente o en el staging de Drive usado por la corrida."},
        {"Paso": 9, "Accion": "Reprocesar", "Detalle": "Ejecuta scripts/run_finmonitor_prod.sh o corre financial_monitor_pipeline.py con --clients y --output."},
        {"Paso": 10, "Accion": "Auditar detalle", "Detalle": f"El archivo objetivo actual es {output_path}. Revisa Razones, QA, Documentos y Auditoria; las hojas raw quedan ocultas para no ensuciar la vista principal."},
    ]
    return pd.DataFrame(rows)


def _status_fill(value):
    normalized = normalize(value)
    if normalized in {"actualizado", "ok", "calculated", "listo para revisar"}:
        return PatternFill("solid", fgColor="E9F7EF")
    if normalized in {"requiere revision", "sin razones", "needs review", "needs_review", "media"}:
        return PatternFill("solid", fgColor="FFF3D6")
    if normalized in {"sin documentos", "missing financial dir", "missing_financial_dir", "no existe carpeta de estados financieros", "alta", "error", "unmapped"}:
        return PatternFill("solid", fgColor="FDE8E8")
    return PatternFill("solid", fgColor="EAF3FF")


def _status_font_color(value):
    normalized = normalize(value)
    if normalized in {"actualizado", "ok", "calculated", "listo para revisar"}:
        return "15803D"
    if normalized in {"requiere revision", "sin razones", "needs review", "needs_review", "media"}:
        return "B45309"
    if normalized in {"sin documentos", "missing financial dir", "missing_financial_dir", "no existe carpeta de estados financieros", "alta", "error", "unmapped"}:
        return "B91C1C"
    return "183A59"


def _dashboard_action(crm, documents, accounts, concepts, ratios, qa):
    missing = pd.DataFrame()
    if not documents.empty and "status" in documents:
        missing = documents[documents["status"].astype(str).eq("missing_financial_dir")]
    if not missing.empty:
        path = _first_nonblank(missing.get("path"))
        return (
            "Corregir ruta fuente",
            "No existe carpeta de estados financieros",
            "Apuntar clients-root/staging a una carpeta valida y reprocesar.",
            path,
        )
    if len(accounts) == 0:
        return (
            "Validar extraccion",
            "No se extrajeron cuentas",
            "Confirmar que los PDFs/Excel sean legibles y que el limite de paginas incluya los estados principales.",
            "",
        )
    if len(concepts) == 0:
        return (
            "Mapear cuentas",
            "No hay conceptos mapeados",
            "Completar Mapping Memory para las etiquetas nuevas y reprocesar.",
            "",
        )
    if len(ratios) == 0:
        return (
            "Revisar covenants",
            "No hay razones calculadas",
            "Validar facility_covenants, periodos y conceptos requeridos para cada razon.",
            "",
        )
    qa_pending = int(qa["status"].eq("needs_review").sum()) if not qa.empty and "status" in qa else 0
    ratio_pending = int(ratios["review_status"].eq("needs_review").sum()) if not ratios.empty and "review_status" in ratios else 0
    if qa_pending or ratio_pending:
        return (
            "Resolver alertas",
            f"{ratio_pending} razones / {qa_pending} QA pendientes",
            "Abrir Razones y QA, filtrar needs_review y documentar el seguimiento.",
            "",
        )
    return (
        "Confirmar seguimiento",
        "Sin alertas principales",
        "Actualizar Responsable, Proxima accion y Fecha actualizacion si aplica.",
        "",
    )


def _set_range_border(ws, cell_range, color="C8D3DF"):
    side = Side(style="thin", color=color)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=side, bottom=side, left=side, right=side)


def style_inicio_dashboard(ws, crm, documents, accounts, concepts, ratios, qa, loan_tape, output_path):
    ws.delete_rows(1, ws.max_row)
    ws.auto_filter.ref = None
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A16"
    ws.sheet_properties.tabColor = "183A59"

    widths = [22, 18, 18, 18, 24, 28, 24, 34]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    client = _first_nonblank(crm["Cliente"]) if not crm.empty and "Cliente" in crm else "Clientes"
    facility = _first_nonblank(crm["Facility"]) if not crm.empty and "Facility" in crm else ""
    status = _first_nonblank(crm["Estatus"]) if not crm.empty and "Estatus" in crm else "Pendiente"
    priority = _first_nonblank(crm["Prioridad"]) if not crm.empty and "Prioridad" in crm else "Media"
    credit = _first_nonblank(crm["Credito"]) if not crm.empty and "Credito" in crm else ""
    product = _first_nonblank(crm["Producto"]) if not crm.empty and "Producto" in crm else ""
    action, blocker, next_step, missing_path = _dashboard_action(crm, documents, accounts, concepts, ratios, qa)
    qa_pending = int(qa["status"].eq("needs_review").sum()) if not qa.empty and "status" in qa else 0
    ratio_pending = int(ratios["review_status"].eq("needs_review").sum()) if not ratios.empty and "review_status" in ratios else 0
    loan_tape_rows = len(loan_tape) if loan_tape is not None else 0

    ws.merge_cells("A1:H1")
    ws["A1"] = "Monitor financiero"
    ws["A1"].font = Font(bold=True, size=18, color="102033")
    ws["A1"].fill = PatternFill("solid", fgColor="E7F6F2")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = f"{client}{' | ' + facility if facility else ''} | Export: {Path(output_path).name}"
    ws["A2"].font = Font(bold=True, color="102033")
    ws["A2"].fill = PatternFill("solid", fgColor="EAF3FF")
    ws.row_dimensions[2].height = 22

    kpis = [
        ("Clientes", len(crm), "Estatus", status, "Razones", len(ratios), "QA pendientes", qa_pending),
        ("Docs indexados", len(documents), "Prioridad", priority, "Razones a revisar", ratio_pending, "Credito", credit),
        ("Cuentas extraidas", len(accounts), "Bloqueo", blocker, "Conceptos", len(concepts), "Producto", product),
    ]
    for row_idx, values in enumerate(kpis, start=4):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx % 2:
                cell.font = Font(bold=True, color="536579")
            else:
                cell.font = Font(bold=True, size=12, color=_status_font_color(value) if col_idx in {4, 8} else "102033")
                cell.number_format = "#,##0"
    _set_range_border(ws, "A4:H6")
    for row in range(4, 7):
        for col in range(1, 9):
            ws.cell(row, col).fill = _status_fill(status) if col in {3, 4} else PatternFill("solid", fgColor="FFFFFF")
        ws.row_dimensions[row].height = 24

    ws.merge_cells("A8:H8")
    ws["A8"] = "Procesos Fin Monitor"
    ws["A8"].font = Font(bold=True, color="102033")
    ws["A8"].fill = PatternFill("solid", fgColor="E7F6F2")
    ws.row_dimensions[8].height = 22

    process_headers = ["Proceso", "Linea", "Incluye", "", "Pantalla / hoja", "", "Objetivo", ""]
    process_fills = ["EAF3FF", "E7F6F2", "FFF3D6", "FDE8E8", "EEF2FF"]
    for col_idx, value in enumerate(process_headers, start=1):
        cell = ws.cell(9, col_idx)
        cell.value = value
        cell.font = Font(bold=True, color="102033")
        cell.fill = PatternFill("solid", fgColor="EAF3FF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_idx, process in enumerate(FIN_MONITOR_PROCESSES, start=10):
        ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=4)
        ws.merge_cells(start_row=row_idx, start_column=5, end_row=row_idx, end_column=6)
        ws.merge_cells(start_row=row_idx, start_column=7, end_row=row_idx, end_column=8)
        values = [
            process["process"],
            process["line"],
            process["includes"],
            process["app_view"],
            process["objective"],
        ]
        target_cols = [1, 2, 3, 5, 7]
        for value, col_idx in zip(values, target_cols):
            cell = ws.cell(row_idx, col_idx)
            cell.value = value
            cell.fill = PatternFill("solid", fgColor=process_fills[row_idx - 10])
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if col_idx == 1:
                cell.font = Font(bold=True, color="102033")
        for col_idx in range(1, 9):
            ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=process_fills[row_idx - 10])
        ws.row_dimensions[row_idx].height = 42
    _set_range_border(ws, "A9:H14")

    headers = ["Siguiente accion", "Cliente", "Estatus", "Prioridad", "Bloqueo principal", "Accion sugerida", "Responsable", "Proxima accion"]
    details = [
        action,
        client,
        status,
        priority,
        blocker,
        next_step,
        _first_nonblank(crm["Responsable"]) if not crm.empty and "Responsable" in crm else "",
        _first_nonblank(crm["Proxima accion"]) if not crm.empty and "Proxima accion" in crm else "",
    ]
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(16, col_idx)
        cell.value = value
        cell.font = Font(bold=True, color="102033")
        cell.fill = PatternFill("solid", fgColor="EAF3FF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for col_idx, value in enumerate(details, start=1):
        cell = ws.cell(17, col_idx)
        cell.value = value
        cell.fill = _status_fill(status)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if col_idx in {3, 4}:
            cell.font = Font(bold=True, color=_status_font_color(value))
    _set_range_border(ws, "A16:H17")
    ws.row_dimensions[16].height = 22
    ws.row_dimensions[17].height = 48

    ws.merge_cells("A20:H20")
    ws["A20"] = "Diagnostico de la corrida"
    ws["A20"].font = Font(bold=True, color="102033")
    ws["A20"].fill = PatternFill("solid", fgColor="E7F6F2")
    checks = [
        ("Senal", "Resultado", "Que significa", "Accion"),
        ("Documentos", len(documents), "Archivos financieros localizados.", "Revisar Documentos para fuente/periodo."),
        ("Extraccion", len(accounts), "Cuentas leidas desde PDF/Excel.", "Validar PDFs/Excel si el conteo es cero."),
        ("Conceptos", len(concepts), "Cuentas normalizadas para calculo.", "Actualizar Mapping Memory si falta mapeo."),
        ("Razones", len(ratios), "Razones financieras calculadas.", "Filtrar Razones por needs_review."),
        ("QA", qa_pending, "Checks que requieren revision.", "Abrir QA y resolver diferencias."),
        ("Loan tape", loan_tape_rows, "Cartera por cliente basada en proxy contable.", "Abrir Loan Tape Cliente y validar contra tape granular."),
    ]
    for row_offset, values in enumerate(checks, start=21):
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_offset, col_idx)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if row_offset == 21:
                cell.font = Font(bold=True, color="102033")
                cell.fill = PatternFill("solid", fgColor="EAF3FF")
            elif col_idx == 1:
                cell.font = Font(bold=True, color="102033")
    _set_range_border(ws, f"A21:D{20 + len(checks)}")
    for row in range(22, 21 + len(checks)):
        ws.row_dimensions[row].height = 28
        ws.cell(row, 2).number_format = "#,##0"

    if missing_path:
        ws.merge_cells("A29:H30")
        ws["A29"] = f"Ruta faltante: {missing_path}"
        ws["A29"].font = Font(bold=True, color="B91C1C")
        ws["A29"].fill = PatternFill("solid", fgColor="FDE8E8")
        ws["A29"].alignment = Alignment(wrap_text=True, vertical="center")
        _set_range_border(ws, "A29:H30", color="F5B7B1")


def style_workbook(writer):
    header_fill = PatternFill("solid", fgColor="16324F")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="0F172A")
    panel_fill = PatternFill("solid", fgColor="EAF2F8")
    border = Border(bottom=Side(style="thin", color="C9D6E2"))
    status_fills = {
        "ok": PatternFill("solid", fgColor="B7E4C7"),
        "ok proxy": PatternFill("solid", fgColor="B7E4C7"),
        "calculated": PatternFill("solid", fgColor="B7E4C7"),
        "extracted": PatternFill("solid", fgColor="B7E4C7"),
        "actualizado": PatternFill("solid", fgColor="B7E4C7"),
        "listo para revisar": PatternFill("solid", fgColor="B7E4C7"),
        "needs_review": PatternFill("solid", fgColor="FCD34D"),
        "requiere revision": PatternFill("solid", fgColor="FCD34D"),
        "revisar": PatternFill("solid", fgColor="FCD34D"),
        "sin razones": PatternFill("solid", fgColor="FCD34D"),
        "sin cartera": PatternFill("solid", fgColor="FCA5A5"),
        "unmapped": PatternFill("solid", fgColor="FCA5A5"),
        "error": PatternFill("solid", fgColor="FCA5A5"),
        "sin documentos": PatternFill("solid", fgColor="FCA5A5"),
        "alta": PatternFill("solid", fgColor="FCA5A5"),
        "media": PatternFill("solid", fgColor="FCD34D"),
        "baja": PatternFill("solid", fgColor="B7E4C7"),
    }
    status_font_colors = {
        "ok": "064E3B",
        "ok proxy": "064E3B",
        "calculated": "064E3B",
        "extracted": "064E3B",
        "actualizado": "064E3B",
        "listo para revisar": "064E3B",
        "needs_review": "7C2D12",
        "requiere revision": "7C2D12",
        "revisar": "7C2D12",
        "sin razones": "7C2D12",
        "sin cartera": "7F1D1D",
        "unmapped": "7F1D1D",
        "error": "7F1D1D",
        "sin documentos": "7F1D1D",
        "alta": "7F1D1D",
        "media": "7C2D12",
        "baja": "064E3B",
    }
    status_headers = {
        "review_status",
        "status",
        "mapping_status",
        "source_method",
        "estatus_crm",
        "prioridad",
        "Estatus",
        "Prioridad",
        "Calidad cartera",
    }
    long_headers = {
        "path",
        "source_ref",
        "raw_value",
        "review_notes",
        "mapping_notes",
        "details",
        "contrato_drive_path",
        "contrato_drive_link",
        "proxima_accion",
        "detalle",
        "Link contrato",
        "Proxima accion",
        "Detalle",
        "Bloqueo principal",
        "Accion sugerida",
        "Notas seguimiento",
        "Notas",
        "notes",
        "formula",
        "Base cartera",
        "Alertas cartera",
        "Base analisis",
        "Fuente arrendamiento",
        "Fuente factoraje",
        "Fuente estimacion",
        "Fuente ultimo periodo",
        "Etiqueta fuente",
        "Formula",
    }
    whole_number_headers = {
        "Docs",
        "Razones revisar",
        "QA revisar",
        "Valor",
        "count",
        "file_size_bytes",
        "unit_scale",
        "numerator",
        "denominator",
        "difference",
        "Cartera neta",
        "Cartera arrendamiento",
        "Cartera factoraje",
        "Cartera bruta",
        "Estimacion preventiva reportada",
        "Ajuste estimacion usado",
        "Cartera neta analizada",
    }
    for ws in writer.book.worksheets:
        if ws.max_row < 1:
            continue
        is_front_sheet = ws.title in FRONT_SHEETS
        ws.sheet_view.showGridLines = not is_front_sheet
        ws.freeze_panes = "A2"
        if ws.title in HIDDEN_SHEETS:
            ws.sheet_state = "hidden"
        if ws.title in FILTERABLE_FRONT_SHEETS and ws.max_row >= 2 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        elif ws.title not in HIDDEN_SHEETS and ws.max_row >= 2 and ws.max_row <= 1000 and ws.max_column <= 30:
            ws.auto_filter.ref = ws.dimensions
        else:
            ws.auto_filter.ref = None
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        headers = [cell.value for cell in ws[1]]
        for col_idx, header in enumerate(headers, start=1):
            letter = get_column_letter(col_idx)
            width = 14
            if header in long_headers:
                width = 44
            elif header in {"filename", "raw_label", "formula", "notas", "Notas", "Notas seguimiento"}:
                width = 32
            elif header in {"client", "cliente", "Cliente", "period", "Periodo", "ratio", "concept", "Concepto", "Concepto normalizado", "Categoria", "Estado", "estatus_crm", "producto_principal", "Estatus", "Producto", "facility_id", "facility_name", "Facility ID", "Facility", "Calidad cartera"}:
                width = 20
            elif header in {"responsable", "Responsable", "prioridad", "Prioridad", "ultimo_periodo", "fecha_actualizacion", "Ultimo periodo", "Fecha actualizacion", "Tipo EEFF"}:
                width = 18
            ws.column_dimensions[letter].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                header = headers[cell.column - 1] if cell.column - 1 < len(headers) else ""
                cell.alignment = Alignment(wrap_text=header in long_headers, vertical="top")
                if header == "result":
                    cell.number_format = '0.0%;[Red](0.0%);-'
                elif header == "Estimacion / cartera bruta":
                    cell.number_format = '0.0%;[Red](0.0%);-'
                elif header == "result_pct":
                    cell.number_format = '0.0;[Red](0.0);-'
                elif header in whole_number_headers or re.match(r"^20\d{2}-\d{2}-\d{2}$", str(header)):
                    cell.number_format = '#,##0;[Red](#,##0);-'
                elif isinstance(cell.value, float):
                    cell.number_format = '#,##0.00;[Red](#,##0.00);-'
                if header in status_headers:
                    fill = status_fills.get(str(cell.value).lower())
                    if fill:
                        cell.fill = fill
                        cell.font = Font(bold=True, color=status_font_colors.get(str(cell.value).lower(), "102033"))
        if not is_front_sheet:
            continue
        if ws.title in {"Inicio", "CRM Clientes", "Actualizar"}:
            for cell in ws[1]:
                cell.fill = title_fill
        if ws.title == "Inicio":
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 56
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, 1).fill = panel_fill
                ws.cell(row_idx, 1).font = Font(bold=True, color="16324F")
                ws.cell(row_idx, 3).number_format = '#,##0'
        if ws.title == "CRM Clientes":
            ws.sheet_view.zoomScale = 90
            ws.sheet_properties.tabColor = "16324F"
            ws.freeze_panes = "B2"
            for col_idx in range(10, 13):
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row_idx, col_idx).number_format = '#,##0'
            if "Cartera neta" in headers:
                col_idx = headers.index("Cartera neta") + 1
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row_idx, col_idx).number_format = '#,##0;[Red](#,##0);-'
            for col_idx in range(6, 9):
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row_idx, col_idx).font = Font(color="0000FF")
                    ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor="FFF2CC")
            if ws.max_row >= 2:
                priority = DataValidation(type="list", formula1='"Alta,Media,Baja"', allow_blank=True)
                status = DataValidation(type="list", formula1='"Actualizado,Listo para revisar,Requiere revision,Sin razones,Sin documentos"', allow_blank=False)
                ws.add_data_validation(priority)
                ws.add_data_validation(status)
                priority.add(f"C2:C{ws.max_row}")
                status.add(f"B2:B{ws.max_row}")
        if ws.title == "Loan Tape Cliente":
            ws.sheet_view.zoomScale = 90
            ws.sheet_properties.tabColor = "0F766E"
            ws.freeze_panes = "C2"
            for row in ws.iter_rows(min_row=2):
                row[11].alignment = Alignment(wrap_text=True, vertical="top")
                row[12].alignment = Alignment(wrap_text=True, vertical="top")
                row[13].alignment = Alignment(wrap_text=True, vertical="top")
                row[14].alignment = Alignment(wrap_text=True, vertical="top")
                row[15].alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[row[0].row].height = 42
        if ws.title == "Estados Financieros":
            ws.sheet_view.zoomScale = 90
            ws.sheet_properties.tabColor = "2563EB"
            ws.freeze_panes = "F2"
            for col_idx, header in enumerate(headers, start=1):
                if re.match(r"^20\d{2}-\d{2}-\d{2}$", str(header)):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
            concept_col = headers.index("Concepto") if "Concepto" in headers else None
            wrap_cols = [headers.index(header) for header in ("Fuente ultimo periodo", "Etiqueta fuente") if header in headers]
            for row in ws.iter_rows(min_row=2):
                if concept_col is not None:
                    row[concept_col].font = Font(bold=True, color="102033")
                for col_idx in wrap_cols:
                    row[col_idx].alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[row[0].row].height = 34
        if ws.title == "Actualizar":
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 100
            for row in ws.iter_rows(min_row=2):
                row[2].alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[row[0].row].height = 36


def export_workbook(path, documents, accounts, concepts, ratios, qa, mapping, credit_evidence, followups=None):
    loan_tape = build_loan_tape_by_client(concepts)
    financial_statements = build_financial_statements_view(concepts)
    crm = build_crm_clients(documents, ratios, qa, followups=followups, loan_tape=loan_tape)
    inicio = build_inicio(crm, documents, accounts, concepts, ratios, qa, loan_tape=loan_tape)
    update_guide = build_update_guide(path)
    audit = pd.DataFrame(
        [
            {"category": "documents", "detail": "indexed", "count": len(documents)},
            {"category": "accounts", "detail": "extracted", "count": len(accounts)},
            {"category": "accounts", "detail": "pdfs_seen", "count": accounts.attrs.get("pdf_count", 0)},
            {"category": "accounts", "detail": "pdf_cache_hits", "count": accounts.attrs.get("cache_hits", 0)},
            {"category": "accounts", "detail": "xlsx_seen", "count": accounts.attrs.get("xlsx_count", 0)},
            {"category": "concepts", "detail": "mapped", "count": len(concepts)},
            {"category": "ratios", "detail": "total", "count": len(ratios)},
            {"category": "ratios", "detail": "calculated", "count": int(ratios["review_status"].eq("calculated").sum()) if not ratios.empty else 0},
            {"category": "ratios", "detail": "needs_review", "count": int(ratios["review_status"].eq("needs_review").sum()) if not ratios.empty else 0},
            {"category": "qa", "detail": "ok", "count": int(qa["status"].eq("ok").sum()) if not qa.empty else 0},
            {"category": "qa", "detail": "needs_review", "count": int(qa["status"].eq("needs_review").sum()) if not qa.empty else 0},
            {"category": "credit_evidence", "detail": "files_found", "count": len(credit_evidence)},
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        inicio.to_excel(writer, sheet_name="Inicio", index=False)
        crm.to_excel(writer, sheet_name="CRM Clientes", index=False)
        update_guide.to_excel(writer, sheet_name="Actualizar", index=False)
        loan_tape.to_excel(writer, sheet_name="Loan Tape Cliente", index=False)
        financial_statements.to_excel(writer, sheet_name="Estados Financieros", index=False)
        ratios.to_excel(writer, sheet_name="Razones", index=False)
        qa.to_excel(writer, sheet_name="QA", index=False)
        documents.to_excel(writer, sheet_name="Documentos", index=False)
        concepts.to_excel(writer, sheet_name="Conceptos", index=False)
        accounts.to_excel(writer, sheet_name="Cuentas Extraidas", index=False)
        credit_evidence.to_excel(writer, sheet_name="Credit Evidence", index=False)
        golden_sample_scaffold(ratios).to_excel(writer, sheet_name="Golden Sample", index=False)
        audit.to_excel(writer, sheet_name="Auditoria", index=False)
        mapping.to_excel(writer, sheet_name="Mapping Memory", index=False)
        style_workbook(writer)
        style_inicio_dashboard(writer.book["Inicio"], crm, documents, accounts, concepts, ratios, qa, loan_tape, path)
    write_crm_sidecar(
        path,
        crm,
        ratios=ratios,
        qa=qa,
        documents=documents,
        concepts=concepts,
        loan_tape=loan_tape,
        financial_statements=financial_statements,
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--profile", choices=["default", "prod"], default="default")
    parser.add_argument("--clients-root", default=str(DEFAULT_CLIENTS_ROOT))
    parser.add_argument("--clients", default="Ventus", help="Comma-separated client folder names")
    parser.add_argument("--client-metadata", default=str(CLIENT_METADATA_PATH))
    parser.add_argument("--client-followups", default=str(CLIENT_FOLLOWUPS_PATH))
    parser.add_argument("--facility", default="", help="Facility id or name to isolate in client metadata and report output.")
    parser.add_argument("--from-period", default="2025-01-01")
    parser.add_argument("--to-period", default="")
    parser.add_argument("--max-documents", type=int, default=40)
    parser.add_argument("--max-pages-per-pdf", type=int, default=5)
    parser.add_argument("--max-file-mb", type=int, default=8)
    parser.add_argument(
        "--pdf-workers",
        type=int,
        default=max(1, min(4, (os.cpu_count() or 2) - 1)),
        help="Parallel PDF extraction workers. Cache hits are still read in-process for fast warm runs.",
    )
    parser.add_argument("--include-undated", action="store_true")
    parser.add_argument(
        "--extra-source",
        action="append",
        default=[],
        help="Additional source file path to include in document discovery. Repeat it or separate paths with '|'.",
    )
    parser.add_argument("--skip-auxiliary-sources", action="store_true", help="Do not auto-include known auxiliary EEFF workbooks.")
    parser.add_argument("--refresh-cache", action="store_true", help="Ignore cached PDF extraction results.")
    parser.add_argument("--skip-credit-evidence", action="store_true", help="Skip broader Drive scan for credit contract evidence.")
    parser.add_argument("--output", default=str(OUTPUT_DIR / "financial_monitor_pipeline.xlsx"))
    args = parser.parse_args()

    clients = [client.strip() for client in args.clients.split(",") if client.strip()]

    if args.profile == "prod":
        default_prod_closed_clients = {normalize(client) for client in clients} == {"ventus"}
        if args.clients_root == str(DEFAULT_CLIENTS_ROOT) and default_prod_closed_clients:
            args.clients_root = (
                "/Users/syscap/Library/CloudStorage/GoogleDrive-rbarron@syscap.com.mx/Shared drives/"
                "Axcess - Crédito y Riesgo/1. Clientes/3. Cerrados, Dormant & Rechazados"
            )
        args.skip_credit_evidence = True
        args.max_pages_per_pdf = min(args.max_pages_per_pdf, 5)
        args.max_file_mb = min(args.max_file_mb, 8)

    started = log_step("starting")
    clients_root = Path(args.clients_root)
    metadata = load_client_metadata(args.client_metadata)
    metadata = filter_client_metadata_by_facility(metadata, args.facility)
    followups = load_client_followups(args.client_followups)
    followups = filter_followups_by_clients(followups, clients)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    step = log_step("discovering documents")
    documents = discover_documents(
        clients_root,
        clients,
        from_period=args.from_period,
        to_period=args.to_period,
        max_documents=args.max_documents,
        metadata=metadata,
        include_undated=args.include_undated,
    )
    auxiliary_sources = [] if args.skip_auxiliary_sources else discover_auxiliary_sources(clients)
    documents = append_extra_sources(documents, clients, [*auxiliary_sources, *args.extra_source], metadata=metadata)
    log_step(f"discovered documents={len(documents)}", step)

    if args.skip_credit_evidence:
        credit_evidence = pd.DataFrame()
    else:
        step = log_step("scanning credit evidence")
        credit_evidence = scan_credit_line_evidence(clients_root, clients)
        log_step(f"credit evidence rows={len(credit_evidence)}", step)

    step = log_step("extracting PDF accounts")
    accounts = extract_accounts(
        documents,
        max_pages=args.max_pages_per_pdf,
        max_file_mb=args.max_file_mb,
        refresh_cache=args.refresh_cache,
        pdf_workers=args.pdf_workers,
    )
    log_step(
        f"extracted accounts={len(accounts)} pdfs={accounts.attrs.get('pdf_count', 0)} xlsx={accounts.attrs.get('xlsx_count', 0)} cache_hits={accounts.attrs.get('cache_hits', 0)}",
        step,
    )

    step = log_step("mapping and calculating")
    mapping = load_mapping_memory()
    mapped_accounts = map_accounts(accounts, mapping)
    concepts = best_concepts(mapped_accounts)
    ratios, qa = calculate_ratios(concepts) if not concepts.empty else (pd.DataFrame(), pd.DataFrame())
    ratios = filter_ratios_by_facility_covenants(ratios, metadata, facility=args.facility)
    pair_qa = pairing_qa(documents)
    if not pair_qa.empty:
        qa = pd.concat([qa, pair_qa], ignore_index=True)
    facility_qa = facility_selection_qa(metadata, facility=args.facility)
    if not facility_qa.empty:
        qa = pd.concat([qa, facility_qa], ignore_index=True)
    log_step(f"mapped concepts={len(concepts)} ratios={len(ratios)}", step)

    step = log_step("exporting workbook")
    export_workbook(Path(args.output), documents, mapped_accounts, concepts, ratios, qa, mapping, credit_evidence, followups=followups)
    log_step(f"exported {args.output}", step)
    log_step("finished", started)
    print(args.output)
    print(f"documents={len(documents)} accounts={len(mapped_accounts)} concepts={len(concepts)} ratios={len(ratios)}")


if __name__ == "__main__":
    main()
