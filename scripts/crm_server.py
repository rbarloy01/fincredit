#!/usr/bin/env python3
import csv
import json
import subprocess
import sys
import time
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "outputs/financial_monitor/financial_monitor_pipeline.xlsx"
CRM_JSON_PATH = WORKBOOK_PATH.with_suffix(".crm.json")
FOLLOWUPS_PATH = ROOT / "config/client_followups.tsv"
RUN_SCRIPT = ROOT / "scripts/run_finmonitor_prod.sh"
FOLLOWUP_COLUMNS = ["client", "responsable", "proxima_accion", "fecha_actualizacion", "seguimiento_notas"]
APP_VERSION = "analysis-workbench-2026-08-26-eeff-ux"
CRM_CACHE = {"mtime_ns": None, "payload": None}
DETAIL_SHEETS = {
    "ratios": "Razones",
    "qa": "QA",
    "documents": "Documentos",
    "concepts": "Conceptos",
    "loanTape": "Loan Tape Cliente",
    "financialStatements": "Estados Financieros",
}


HTML = r"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Financial Monitor</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f5f7fa;
      --sidebar: #101820;
      --sidebar-soft: #18232e;
      --panel: #ffffff;
      --panel-soft: #f8fafc;
      --ink: #17202a;
      --muted: #627084;
      --line: #d8e0ea;
      --line-strong: #b8c4d2;
      --accent: #0f766e;
      --accent-soft: #dff4f1;
      --blue: #1d4ed8;
      --warn: #a16207;
      --bad: #b42318;
      --good: #067647;
      --field: #ffffff;
      --field-edit: #fff8db;
      --shadow: 0 12px 34px rgba(16, 24, 40, 0.08);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    .shell {
      min-height: 100vh;
      display: grid;
      grid-template-columns: 260px minmax(560px, 1fr) 390px;
    }
    aside {
      background: var(--sidebar);
      color: #f8fafc;
      padding: 20px;
      overflow: auto;
    }
    main {
      min-width: 0;
      padding: 18px;
      overflow: auto;
    }
    .detail {
      min-width: 0;
      border-left: 1px solid var(--line);
      background: var(--panel);
      padding: 18px;
      overflow: auto;
    }
    h1, h2, h3, p { margin: 0; }
    .brand {
      display: grid;
      gap: 8px;
      padding-bottom: 18px;
      border-bottom: 1px solid rgba(255,255,255,.16);
    }
    .brand h1 { font-size: 21px; letter-spacing: 0; }
    .stamp { color: #b9c4d0; font-size: 12px; line-height: 1.4; min-height: 18px; }
    .section-title {
      margin: 20px 0 10px;
      color: #b9c4d0;
      font-size: 11px;
      font-weight: 800;
      letter-spacing: .08em;
      text-transform: uppercase;
    }
    .metric {
      display: grid;
      gap: 4px;
      padding: 12px;
      margin-bottom: 8px;
      background: var(--sidebar-soft);
      border: 1px solid rgba(255,255,255,.12);
      border-radius: 8px;
    }
    .metric span { color: #b9c4d0; font-size: 12px; }
    .metric strong { color: #ffffff; font-size: 23px; letter-spacing: 0; }
    .process-list { display: grid; gap: 8px; }
    .process {
      display: grid;
      gap: 3px;
      padding: 10px 11px;
      background: var(--sidebar-soft);
      border-left: 3px solid var(--accent);
      border-radius: 6px;
    }
    .process strong { font-size: 12px; }
    .process span { color: #b9c4d0; font-size: 11px; }
    .side-actions {
      display: grid;
      gap: 8px;
      padding-top: 14px;
      border-top: 1px solid rgba(255,255,255,.16);
      margin-top: 18px;
    }
    .caption { color: #b9c4d0; font-size: 12px; line-height: 1.4; }
    .topbar {
      display: grid;
      grid-template-columns: 1fr auto;
      align-items: end;
      gap: 14px;
      margin-bottom: 14px;
    }
    .topbar h1 { font-size: 24px; letter-spacing: 0; }
    .topbar p { color: var(--muted); font-size: 13px; margin-top: 4px; }
    .summary-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 10px;
      margin-bottom: 14px;
    }
    .summary-card {
      min-height: 86px;
      padding: 13px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
    }
    .summary-card span { display: block; color: var(--muted); font-size: 12px; }
    .summary-card strong { display: block; font-size: 24px; margin-top: 8px; letter-spacing: 0; }
    .summary-card small { display: block; color: var(--muted); font-size: 11px; margin-top: 4px; }
    .toolbar {
      display: grid;
      grid-template-columns: minmax(220px, 1fr) minmax(130px, 150px) minmax(140px, 170px);
      gap: 8px;
      margin-bottom: 10px;
    }
    #sort { grid-column: 1 / -1; }
    input, select, textarea, button { font: inherit; }
    input, select, textarea {
      width: 100%;
      min-height: 36px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: var(--field);
      color: var(--ink);
      padding: 8px 10px;
      outline: none;
    }
    input:focus, select:focus, textarea:focus {
      border-color: var(--accent);
      box-shadow: 0 0 0 3px rgba(15, 118, 110, .13);
    }
    textarea { min-height: 92px; resize: vertical; line-height: 1.35; }
    button {
      min-height: 36px;
      border: 1px solid #0f172a;
      border-radius: 6px;
      background: #0f172a;
      color: #ffffff;
      padding: 8px 12px;
      cursor: pointer;
      font-weight: 700;
      white-space: nowrap;
    }
    button.secondary {
      background: #ffffff;
      color: var(--ink);
      border-color: var(--line-strong);
    }
    button.ghost {
      background: transparent;
      color: #ffffff;
      border-color: rgba(255,255,255,.25);
    }
    button:disabled { cursor: wait; opacity: .65; }
    .table-wrap {
      overflow: auto;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      max-height: calc(100vh - 260px);
    }
    table {
      width: 100%;
      min-width: 1180px;
      border-collapse: collapse;
      font-size: 13px;
    }
    th {
      position: sticky;
      top: 0;
      z-index: 1;
      background: #17202a;
      color: #ffffff;
      text-align: left;
      padding: 9px 10px;
      white-space: nowrap;
    }
    td {
      padding: 9px 10px;
      border-top: 1px solid var(--line);
      vertical-align: middle;
      white-space: nowrap;
    }
    tbody tr { cursor: pointer; }
    tbody tr:hover { background: #edf7f5; }
    tbody tr.selected { background: var(--accent-soft); outline: 2px solid var(--accent); outline-offset: -2px; }
    .num { text-align: right; font-variant-numeric: tabular-nums; }
    .pill {
      display: inline-flex;
      align-items: center;
      min-height: 24px;
      padding: 3px 8px;
      border-radius: 999px;
      font-size: 12px;
      font-weight: 800;
      border: 1px solid transparent;
      white-space: nowrap;
    }
    .pill.bad { background: #fee4e2; color: var(--bad); border-color: #fecdca; }
    .pill.warn { background: #fef0c7; color: var(--warn); border-color: #fedf89; }
    .pill.good { background: #dcfae6; color: var(--good); border-color: #abefc6; }
    .pill.info { background: #dbeafe; color: var(--blue); border-color: #bfdbfe; }
    .client-head {
      display: grid;
      grid-template-columns: minmax(0, 1fr) 128px;
      gap: 10px;
      align-items: start;
      padding-bottom: 12px;
      border-bottom: 1px solid var(--line);
    }
    .client-head h1 { font-size: 22px; letter-spacing: 0; overflow-wrap: anywhere; }
    .client-subtitle { color: var(--muted); font-size: 12px; line-height: 1.4; margin-top: 4px; }
    .snapshot {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
      margin: 12px 0;
    }
    .mini {
      min-height: 76px;
      padding: 10px;
      background: var(--panel-soft);
      border: 1px solid var(--line);
      border-radius: 8px;
    }
    .mini span { display: block; color: var(--muted); font-size: 11px; }
    .mini strong { display: block; margin-top: 6px; font-size: 17px; overflow-wrap: anywhere; }
    .tabs {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 4px;
      margin: 12px 0;
      padding: 4px;
      background: #eef2f6;
      border-radius: 8px;
    }
    .tab {
      min-height: 34px;
      padding: 7px 6px;
      border: 0;
      background: transparent;
      color: var(--muted);
      border-radius: 6px;
      font-size: 12px;
      font-weight: 800;
      overflow: hidden;
      text-overflow: ellipsis;
    }
    .tab.active { background: #ffffff; color: var(--ink); box-shadow: 0 1px 3px rgba(16, 24, 40, .13); }
    .panel-stack { display: grid; gap: 10px; }
    .card {
      background: #ffffff;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
    }
    .card-head {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 10px;
      align-items: start;
      margin-bottom: 8px;
    }
    .card h3 { font-size: 14px; overflow-wrap: anywhere; }
    .card p, .kv, .empty { color: var(--muted); font-size: 12px; line-height: 1.45; }
    .formula {
      margin-top: 8px;
      padding: 9px;
      background: #f8fafc;
      border: 1px solid var(--line);
      border-radius: 6px;
      color: #344054;
      font-size: 12px;
      line-height: 1.45;
      overflow-wrap: anywhere;
    }
    .kv-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
      margin-top: 8px;
    }
    .kv {
      padding: 8px;
      background: var(--panel-soft);
      border-radius: 6px;
    }
    .kv b { display: block; color: var(--ink); font-size: 13px; margin-top: 3px; overflow-wrap: anywhere; }
    .list-table {
      min-width: 0;
      font-size: 12px;
    }
    .list-table th { position: static; padding: 8px; background: #eef2f6; color: var(--ink); }
    .list-table td { padding: 8px; white-space: normal; overflow-wrap: anywhere; }
    .statement-table td:nth-child(3),
    .statement-table td:nth-child(4) {
      text-align: right;
      font-variant-numeric: tabular-nums;
      white-space: nowrap;
    }
    .statement-table tr.section td {
      background: #f1f5f9;
      color: #334155;
      font-weight: 800;
    }
    .period-strip {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(92px, 1fr));
      gap: 6px;
      margin: 10px 0;
    }
    .period-chip {
      min-height: 34px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #ffffff;
      color: var(--ink);
      font-size: 12px;
      font-weight: 800;
    }
    .period-chip.active {
      border-color: var(--accent);
      background: var(--accent-soft);
      color: #134e4a;
    }
    .source-note {
      color: var(--muted);
      font-size: 11px;
      line-height: 1.4;
      overflow-wrap: anywhere;
    }
    .fields { display: grid; gap: 10px; }
    label { display: grid; gap: 5px; color: var(--muted); font-size: 12px; }
    label span { color: var(--ink); font-size: 13px; font-weight: 800; }
    .edit { background: var(--field-edit); }
    .statusline {
      min-height: 18px;
      margin-top: 9px;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.35;
      white-space: pre-wrap;
    }
    .hide { display: none; }
    @media (max-width: 1180px) {
      .shell { grid-template-columns: 1fr; }
      aside, .detail { border: 0; }
      .summary-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .toolbar { grid-template-columns: 1fr 1fr; }
      .table-wrap { max-height: 58vh; }
    }
    @media (max-width: 640px) {
      main, aside, .detail { padding: 14px; }
      .summary-grid, .snapshot, .kv-grid, .toolbar, .client-head { grid-template-columns: 1fr; }
      .tabs { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    }
  </style>
</head>
<body>
  <div class="shell">
    <aside>
      <div class="brand">
        <h1>Financial Monitor</h1>
        <div id="stamp" class="stamp"></div>
      </div>
      <div class="section-title">Resumen ejecutivo</div>
      <div id="metrics"></div>
      <div class="section-title">Flujo</div>
      <div id="processes" class="process-list"></div>
      <div class="side-actions">
        <button id="run" type="button">Reprocesar datos</button>
        <button id="openXlsx" class="ghost" type="button">Exportar Excel</button>
        <div class="caption">Excel queda como respaldo completo; el analisis diario vive aqui.</div>
        <div id="runStatus" class="statusline"></div>
      </div>
    </aside>
    <main>
      <div class="topbar">
        <div>
          <h1>Clientes y covenants</h1>
          <p id="mainSubtitle">Vista viva de cartera, razones, QA y trazabilidad.</p>
        </div>
        <button id="refresh" class="secondary" type="button">Actualizar</button>
      </div>
      <div id="summaryCards" class="summary-grid"></div>
      <div class="toolbar">
        <input id="search" type="search" placeholder="Buscar cliente, bloqueo, accion o producto" />
        <select id="priority">
          <option value="">Prioridad</option>
          <option>Alta</option>
          <option>Media</option>
          <option>Baja</option>
        </select>
        <select id="status">
          <option value="">Estatus</option>
          <option>Requiere revision</option>
          <option>Sin documentos</option>
          <option>Sin razones</option>
          <option>Actualizado</option>
          <option>Listo para revisar</option>
        </select>
        <select id="sort">
          <option value="priority">Orden: prioridad</option>
          <option value="client">Orden: cliente</option>
          <option value="portfolio">Orden: cartera</option>
          <option value="period">Orden: periodo</option>
        </select>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>Cliente</th><th>Estatus</th><th>Prioridad</th><th>Bloqueo</th><th>Accion</th>
              <th>Responsable</th><th>Periodo</th><th class="num">Cartera</th>
              <th>Calidad</th><th class="num">Razones</th><th class="num">QA</th><th class="num">Docs</th><th>Producto</th>
            </tr>
          </thead>
          <tbody id="rows"></tbody>
        </table>
      </div>
    </main>
    <section class="detail">
      <div class="client-head">
        <div>
          <h1 id="clientTitle">Cliente</h1>
          <div id="clientSubtitle" class="client-subtitle"></div>
        </div>
        <label><span>Periodo</span><select id="periodSelect"></select></label>
      </div>
      <div id="snapshot" class="snapshot"></div>
      <div class="tabs">
        <button class="tab active" data-tab="financials" type="button">EEFF</button>
        <button class="tab" data-tab="analysis" type="button">Razones</button>
        <button class="tab" data-tab="qa" type="button">QA</button>
        <button class="tab" data-tab="documents" type="button">Docs</button>
        <button class="tab" data-tab="extraction" type="button">Extraccion</button>
        <button class="tab" data-tab="followup" type="button">Seguimiento</button>
      </div>
      <div id="detailPanel" class="panel-stack"></div>
      <div id="saveStatus" class="statusline"></div>
    </section>
  </div>
  <script>
    const processes = [
      ["Dashboard", "KPIs y foco ejecutivo"],
      ["CRM", "Responsables y siguiente accion"],
      ["EEFF", "BG/ER por periodo y fuente"],
      ["Analisis", "Razones y formulas"],
      ["Quality", "QA, documentos y fuentes"],
      ["Export", "Workbook de auditoria"],
    ];
    const state = {
      rows: [], ratios: [], qa: [], documents: [], concepts: [], loanTape: [], financialStatements: [],
      selected: null, tab: "financials", period: "", renderTimer: null,
    };
    const byId = (id) => document.getElementById(id);
    const clean = (value) => value == null ? "" : String(value);
    const h = (value) => clean(value).replace(/[&<>"']/g, (char) => ({ "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;", "'": "&#39;" }[char]));
    const num = (value) => value === "" || value == null ? "" : Number(value || 0).toLocaleString("en-US");
    const money = (value) => value === "" || value == null ? "" : Number(value || 0).toLocaleString("en-US", { maximumFractionDigits: 0 });
    const pct = (value) => value === "" || value == null ? "" : `${Number(value || 0).toLocaleString("en-US", { maximumFractionDigits: 1 })}%`;
    const rowClient = (row) => clean(row.client || row.Cliente);
    const rowPeriod = (row) => clean(row.period || row.Periodo);
    const periodKeys = (row) => Object.keys(row || {}).filter((key) => /^20\d{2}-\d{2}-\d{2}$/.test(key));
    const financialRowsForClient = () => clientRows(state.financialStatements);
    const financialPeriodsForSelected = () => {
      const periods = new Set();
      financialRowsForClient().forEach((row) => {
        periodKeys(row).forEach((period) => {
          const value = row[period];
          if (value !== "" && value != null) periods.add(period);
        });
      });
      return [...periods].sort();
    };
    const statusClass = (value) => {
      const text = clean(value).toLowerCase();
      if (text.includes("review") || text.includes("revision") || text.includes("alta") || text.includes("sin cartera") || text.includes("unmapped")) return "bad";
      if (text.includes("media") || text.includes("sin razones") || text.includes("proxy")) return "warn";
      if (text.includes("ok") || text.includes("calculated") || text.includes("actualizado") || text.includes("listo") || text.includes("baja")) return "good";
      return "info";
    };
    const pill = (value) => `<span class="pill ${statusClass(value)}">${h(value)}</span>`;

    async function api(path, options) {
      const response = await fetch(path, options);
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || response.statusText);
      return data;
    }

    function metric(label, value) {
      return `<div class="metric"><span>${h(label)}</span><strong>${num(value)}</strong></div>`;
    }

    function summaryCard(label, value, detail) {
      return `<div class="summary-card"><span>${h(label)}</span><strong>${h(value)}</strong><small>${h(detail)}</small></div>`;
    }

    function renderMetrics(summary) {
      byId("metrics").innerHTML = [
        metric("Clientes", summary.clients),
        metric("Prioridad alta", summary.high),
        metric("Razones revisar", summary.ratioReview),
        metric("QA revisar", summary.qaReview),
      ].join("");
      byId("summaryCards").innerHTML = [
        summaryCard("Cartera neta", money(summary.netPortfolio), "Proxy contable / loan tape"),
        summaryCard("Periodos EEFF", num(summary.financialStatementPeriods), "BG/ER visibles"),
        summaryCard("Documentos", num(summary.documents), "Fuentes indexadas"),
        summaryCard("Razones OK", `${num(summary.calculatedRatios)} / ${num(summary.ratios)}`, "Calculos listos"),
      ].join("");
    }

    function renderProcesses() {
      byId("processes").innerHTML = processes.map(([name, detail]) => `
        <div class="process"><strong>${h(name)}</strong><span>${h(detail)}</span></div>
      `).join("");
    }

    function priorityRank(row) {
      return { Alta: 0, Media: 1, Baja: 2 }[row["Prioridad"]] ?? 3;
    }

    function filteredRows() {
      const q = byId("search").value.trim().toLowerCase();
      const priority = byId("priority").value;
      const status = byId("status").value;
      const sort = byId("sort").value;
      const rows = state.rows.filter((row) => {
        if (priority && row["Prioridad"] !== priority) return false;
        if (status && row["Estatus"] !== status) return false;
        if (!q) return true;
        return ["Cliente", "Bloqueo principal", "Accion sugerida", "Responsable", "Proxima accion", "Producto", "Calidad cartera", "Alertas cartera", "Base cartera"]
          .some((key) => clean(row[key]).toLowerCase().includes(q));
      });
      return rows.sort((a, b) => {
        if (sort === "client") return clean(a["Cliente"]).localeCompare(clean(b["Cliente"]));
        if (sort === "portfolio") return Number(b["Cartera neta"] || 0) - Number(a["Cartera neta"] || 0);
        if (sort === "period") return clean(b["Ultimo periodo"]).localeCompare(clean(a["Ultimo periodo"]));
        return priorityRank(a) - priorityRank(b) || Number(b["QA revisar"] || 0) - Number(a["QA revisar"] || 0);
      });
    }

    function renderRows() {
      const rows = filteredRows();
      byId("rows").innerHTML = rows.map((row) => {
        const selected = state.selected && state.selected["Cliente"] === row["Cliente"] ? " selected" : "";
        return `<tr class="${selected}" data-client="${h(row["Cliente"])}">
          <td><strong>${h(row["Cliente"])}</strong></td>
          <td>${pill(row["Estatus"])}</td>
          <td>${pill(row["Prioridad"])}</td>
          <td>${h(row["Bloqueo principal"])}</td>
          <td>${h(row["Accion sugerida"])}</td>
          <td>${h(row["Responsable"])}</td>
          <td>${h(row["Ultimo periodo"])}</td>
          <td class="num">${money(row["Cartera neta"])}</td>
          <td>${pill(row["Calidad cartera"])}</td>
          <td class="num">${num(row["Razones revisar"])}</td>
          <td class="num">${num(row["QA revisar"])}</td>
          <td class="num">${num(row["Docs"])}</td>
          <td>${h(row["Producto"])}</td>
        </tr>`;
      }).join("");
      byId("mainSubtitle").textContent = `${rows.length} clientes visibles`;
    }

    function clientRows(collection) {
      const client = clean(state.selected?.["Cliente"]);
      return collection.filter((row) => rowClient(row) === client);
    }

    function periodRows(collection) {
      const period = state.period;
      return clientRows(collection).filter((row) => !period || rowPeriod(row) === period);
    }

    function selectedLoanTape() {
      return clientRows(state.loanTape).find((row) => rowPeriod(row) === state.period) || clientRows(state.loanTape).slice(-1)[0] || {};
    }

    function periodsForSelected() {
      const values = new Set();
      [state.ratios, state.qa, state.documents, state.concepts, state.loanTape].forEach((collection) => {
        clientRows(collection).forEach((row) => { if (rowPeriod(row)) values.add(rowPeriod(row)); });
      });
      financialPeriodsForSelected().forEach((period) => values.add(period));
      return [...values].sort();
    }

    function renderPeriodSelect() {
      const periods = periodsForSelected();
      if (!state.period || !periods.includes(state.period)) state.period = periods[periods.length - 1] || "";
      byId("periodSelect").innerHTML = periods.length
        ? periods.map((period) => `<option ${period === state.period ? "selected" : ""}>${h(period)}</option>`).join("")
        : `<option value="">Sin periodo</option>`;
    }

    function renderSnapshot() {
      const row = state.selected || {};
      const ratios = periodRows(state.ratios);
      const qa = periodRows(state.qa);
      const docs = periodRows(state.documents);
      const loan = selectedLoanTape();
      const financials = financialRowsForClient().filter((row) => state.period && row[state.period] !== "" && row[state.period] != null);
      byId("clientTitle").textContent = clean(row["Cliente"] || "Cliente");
      byId("clientSubtitle").textContent = `${clean(row["Estatus"])} / ${clean(row["Prioridad"])} / ${clean(row["Facility"] || row["Facility ID"])}`;
      byId("snapshot").innerHTML = [
        mini("Bloqueo", row["Bloqueo principal"] || "Listo"),
        mini("Cartera neta", money(loan["Cartera neta analizada"] || row["Cartera neta"])),
        mini("Conceptos EEFF", financials.length),
        mini("Razones", `${ratios.filter((item) => item.review_status === "needs_review").length} revisar / ${ratios.length}`),
        mini("QA", `${qa.filter((item) => item.status === "needs_review").length} revisar / ${qa.length}`),
        mini("Docs", docs.length),
      ].join("");
    }

    function mini(label, value) {
      return `<div class="mini"><span>${h(label)}</span><strong>${h(value)}</strong></div>`;
    }

    function renderTabs() {
      document.querySelectorAll(".tab").forEach((tab) => tab.classList.toggle("active", tab.dataset.tab === state.tab));
    }

    function ratioCard(row) {
      const result = row.result_pct !== "" && row.result_pct != null ? pct(row.result_pct) : "";
      return `<div class="card">
        <div class="card-head">
          <h3>${h(row.ratio)}</h3>
          ${pill(row.review_status)}
        </div>
        <div class="kv-grid">
          <div class="kv">Resultado<b>${h(result || "Sin calculo")}</b></div>
          <div class="kv">Periodo<b>${h(row.period)}</b></div>
          <div class="kv">Numerador<b>${money(row.numerator)}</b></div>
          <div class="kv">Denominador<b>${money(row.denominator)}</b></div>
        </div>
        <div class="formula">${h(row.formula)}${row.review_notes ? `<br>${h(row.review_notes)}` : ""}</div>
      </div>`;
    }

    function renderFinancials() {
      const periods = financialPeriodsForSelected();
      const selected = state.period || periods[periods.length - 1] || "";
      const rows = financialRowsForClient();
      if (!rows.length) return `<div class="empty">No hay estados financieros mapeados para este cliente.</div>`;
      const periodButtons = periods.map((period) => `
        <button class="period-chip ${period === selected ? "active" : ""}" data-period="${h(period)}" type="button">${h(period.slice(0, 7))}</button>
      `).join("");
      const visible = rows.filter((row) => row[selected] !== "" && row[selected] != null);
      const latestSource = visible.find((row) => row["Fuente ultimo periodo"]) || {};
      let currentSection = "";
      const body = visible.map((row) => {
        const section = `${row.Estado} / ${row.Categoria}`;
        const sectionRow = section !== currentSection
          ? `<tr class="section"><td colspan="5">${h(section)}</td></tr>`
          : "";
        currentSection = section;
        return `${sectionRow}<tr>
          <td>${h(row.Concepto)}</td>
          <td>${h(row["Concepto normalizado"])}</td>
          <td>${money(row[selected])}</td>
          <td>${h(row.Mapping)}</td>
          <td class="source-note">${h(row["Etiqueta fuente"] || row["Fuente ultimo periodo"])}</td>
        </tr>`;
      }).join("");
      return `
        <div class="card">
          <div class="card-head">
            <h3>Estados financieros</h3>
            ${pill(visible.length ? "Mapeado" : "Sin datos")}
          </div>
          <div class="period-strip">${periodButtons}</div>
          <div class="kv-grid">
            <div class="kv">Periodo seleccionado<b>${h(selected)}</b></div>
            <div class="kv">Conceptos visibles<b>${num(visible.length)}</b></div>
            <div class="kv">Documentos del periodo<b>${num(periodRows(state.documents).length)}</b></div>
            <div class="kv">Fuente reciente<b>${h(latestSource["Fuente ultimo periodo"] || "")}</b></div>
          </div>
        </div>
        <div class="card"><table class="list-table statement-table">
          <thead><tr><th>Concepto</th><th>Mapping</th><th>${h(selected)}</th><th>Estatus</th><th>Fuente</th></tr></thead>
          <tbody>${body || `<tr><td colspan="5">Sin conceptos para este periodo.</td></tr>`}</tbody>
        </table></div>
      `;
    }

    function renderAnalysis() {
      const ratios = periodRows(state.ratios);
      const loan = selectedLoanTape();
      const ratioHtml = ratios.length ? ratios.map(ratioCard).join("") : `<div class="empty">No hay razones para este periodo.</div>`;
      return `
        <div class="card">
          <div class="card-head"><h3>Cartera analizada</h3>${pill(loan["Calidad cartera"] || state.selected?.["Calidad cartera"])}</div>
          <div class="kv-grid">
            <div class="kv">Arrendamiento<b>${money(loan["Cartera arrendamiento"])}</b></div>
            <div class="kv">Factoraje<b>${money(loan["Cartera factoraje"])}</b></div>
            <div class="kv">Estimacion usada<b>${money(loan["Ajuste estimacion usado"])}</b></div>
            <div class="kv">Neta<b>${money(loan["Cartera neta analizada"] || state.selected?.["Cartera neta"])}</b></div>
          </div>
          <div class="formula">${h(loan["Formula"] || "")}<br>${h(loan["Alertas cartera"] || state.selected?.["Alertas cartera"] || "")}</div>
        </div>
        ${ratioHtml}
      `;
    }

    function renderQA() {
      const rows = periodRows(state.qa);
      if (!rows.length) return `<div class="empty">No hay checks QA para este periodo.</div>`;
      return rows.map((row) => `<div class="card">
        <div class="card-head"><h3>${h(row.check)}</h3>${pill(row.status)}</div>
        <div class="kv-grid">
          <div class="kv">Diferencia<b>${money(row.difference)}</b></div>
          <div class="kv">Periodo<b>${h(row.period)}</b></div>
        </div>
        <div class="formula">${h(row.details)}</div>
      </div>`).join("");
    }

    function renderDocuments() {
      const rows = periodRows(state.documents);
      if (!rows.length) return `<div class="empty">No hay documentos para este periodo.</div>`;
      return `<div class="card"><table class="list-table">
        <thead><tr><th>Archivo</th><th>Tipo</th><th>Calidad</th><th>Estatus</th></tr></thead>
        <tbody>${rows.map((row) => `<tr>
          <td>${h(row.filename)}</td><td>${h(row.statement)}</td><td>${h(row.source_quality)}</td><td>${pill(row.status)}</td>
        </tr>`).join("")}</tbody>
      </table></div>`;
    }

    function renderExtraction() {
      const mapped = periodRows(state.concepts).slice(0, 80);
      if (!mapped.length) return `<div class="empty">No hay conceptos mapeados para este periodo.</div>`;
      return `<div class="card"><table class="list-table">
        <thead><tr><th>Concepto</th><th>Cuenta leida</th><th class="num">Valor</th><th>Fuente</th></tr></thead>
        <tbody>${mapped.map((row) => `<tr>
          <td><strong>${h(row.concept)}</strong></td>
          <td>${h(row.raw_label)}</td>
          <td class="num">${money(row.value)}</td>
          <td>${h(row.filename || row.source_ref)}</td>
        </tr>`).join("")}</tbody>
      </table></div>`;
    }

    function renderFollowup() {
      const row = state.selected || {};
      const link = clean(row["Link contrato"]);
      return `<div class="card">
        <div class="kv-grid">
          <div class="kv">Accion sugerida<b>${h(row["Accion sugerida"])}</b></div>
          <div class="kv">Contrato<b>${link ? `<a href="${h(link)}" target="_blank" rel="noreferrer">Abrir</a>` : ""}</b></div>
        </div>
      </div>
      <div class="card fields">
        <label><span>Responsable</span><input id="owner" class="edit" value="${h(row["Responsable"])}" /></label>
        <label><span>Proxima accion</span><textarea id="nextAction" class="edit">${h(row["Proxima accion"])}</textarea></label>
        <label><span>Fecha actualizacion</span><input id="updateDate" class="edit" type="date" value="${h(clean(row["Fecha actualizacion"]).slice(0, 10))}" /></label>
        <label><span>Notas seguimiento</span><textarea id="notes" class="edit">${h(row["Notas seguimiento"])}</textarea></label>
        <button id="save" type="button">Guardar seguimiento</button>
      </div>`;
    }

    function renderDetail() {
      if (!state.selected) {
        byId("detailPanel").innerHTML = `<div class="empty">Sin cliente seleccionado.</div>`;
        return;
      }
      renderPeriodSelect();
      renderSnapshot();
      renderTabs();
      const renderers = {
        financials: renderFinancials,
        analysis: renderAnalysis,
        qa: renderQA,
        documents: renderDocuments,
        extraction: renderExtraction,
        followup: renderFollowup,
      };
      byId("detailPanel").innerHTML = renderers[state.tab]();
      const saveButton = byId("save");
      if (saveButton) saveButton.addEventListener("click", save);
    }

    function selectClient(client) {
      const row = state.rows.find((item) => item["Cliente"] === client) || state.rows[0];
      state.selected = row || null;
      state.period = "";
      byId("saveStatus").textContent = "";
      renderRows();
      renderDetail();
    }

    async function load() {
      byId("stamp").textContent = "Cargando...";
      const data = await api("/api/crm");
      Object.assign(state, {
        rows: data.rows || [],
        ratios: data.ratios || [],
        qa: data.qa || [],
        documents: data.documents || [],
        concepts: data.concepts || [],
        loanTape: data.loanTape || [],
        financialStatements: data.financialStatements || [],
      });
      renderMetrics(data.summary || {});
      byId("stamp").textContent = data.workbookUpdated ? `Actualizado: ${data.workbookUpdated}` : "";
      renderRows();
      selectClient(state.selected?.["Cliente"] || state.rows[0]?.["Cliente"]);
    }

    async function save() {
      if (!state.selected) return;
      const saveButton = byId("save");
      saveButton.disabled = true;
      byId("saveStatus").textContent = "Guardando...";
      try {
        const payload = {
          client: state.selected["Cliente"],
          responsable: byId("owner").value,
          proxima_accion: byId("nextAction").value,
          fecha_actualizacion: byId("updateDate").value,
          seguimiento_notas: byId("notes").value,
        };
        await api("/api/followups", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload),
        });
        Object.assign(state.selected, {
          "Responsable": payload.responsable,
          "Proxima accion": payload.proxima_accion,
          "Fecha actualizacion": payload.fecha_actualizacion,
          "Notas seguimiento": payload.seguimiento_notas,
        });
        byId("saveStatus").textContent = "Guardado.";
        renderRows();
      } catch (error) {
        byId("saveStatus").textContent = error.message;
      } finally {
        saveButton.disabled = false;
      }
    }

    async function runPipeline() {
      byId("run").disabled = true;
      byId("runStatus").textContent = "Procesando...";
      try {
        const data = await api("/api/run", { method: "POST" });
        byId("runStatus").textContent = data.message;
        await load();
      } catch (error) {
        byId("runStatus").textContent = error.message;
      } finally {
        byId("run").disabled = false;
      }
    }

    function scheduleRenderRows() {
      clearTimeout(state.renderTimer);
      state.renderTimer = setTimeout(renderRows, 80);
    }

    byId("rows").addEventListener("click", (event) => {
      const row = event.target.closest("tr[data-client]");
      if (row) selectClient(row.dataset.client);
    });
    byId("detailPanel").addEventListener("click", (event) => {
      const chip = event.target.closest("[data-period]");
      if (!chip) return;
      state.period = chip.dataset.period;
      renderDetail();
    });
    byId("search").addEventListener("input", scheduleRenderRows);
    byId("priority").addEventListener("change", renderRows);
    byId("status").addEventListener("change", renderRows);
    byId("sort").addEventListener("change", renderRows);
    byId("refresh").addEventListener("click", load);
    byId("run").addEventListener("click", runPipeline);
    byId("openXlsx").addEventListener("click", () => api("/api/open-workbook", { method: "POST" }));
    byId("periodSelect").addEventListener("change", (event) => { state.period = event.target.value; renderDetail(); });
    document.querySelectorAll(".tab").forEach((tab) => tab.addEventListener("click", () => { state.tab = tab.dataset.tab; renderDetail(); }));
    renderProcesses();
    load().catch((error) => byId("stamp").textContent = error.message);
  </script>
</body>
</html>
"""


def _json_response(handler, status, payload):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def _read_body(handler):
    length = int(handler.headers.get("Content-Length", "0") or "0")
    raw = handler.rfile.read(length).decode("utf-8") if length else "{}"
    return json.loads(raw)


def _cell_value(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()[:10]
    return value


def read_crm_rows():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")
    workbook_mtime_ns = WORKBOOK_PATH.stat().st_mtime_ns
    if CRM_JSON_PATH.exists():
        try:
            payload = json.loads(CRM_JSON_PATH.read_text(encoding="utf-8"))
            if payload.get("source_mtime_ns") == workbook_mtime_ns:
                return payload.get("rows", [])
        except (OSError, json.JSONDecodeError):
            pass
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
      sheet = workbook["CRM Clientes"]
      headers = [_cell_value(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
      rows = []
      for row in sheet.iter_rows(min_row=2, values_only=True):
          record = {headers[index]: _cell_value(value) for index, value in enumerate(row) if index < len(headers)}
          if record.get("Cliente"):
              rows.append(record)
      return rows
    finally:
      workbook.close()


def read_sheet_records(workbook, sheet_name, max_rows=None):
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    headers = [_cell_value(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        if max_rows and index > max_rows:
            break
        record = {headers[col_index]: _cell_value(value) for col_index, value in enumerate(row) if col_index < len(headers)}
        if any(str(value).strip() for value in record.values() if value is not None):
            rows.append(record)
    return rows


def read_workbook_payload():
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        payload = {
            "rows": read_sheet_records(workbook, "CRM Clientes"),
        }
        for key, sheet_name in DETAIL_SHEETS.items():
            payload[key] = read_sheet_records(workbook, sheet_name, max_rows=5000 if key == "concepts" else None)
        return payload
    finally:
        workbook.close()


def read_sidecar_payload(workbook_mtime_ns):
    if not CRM_JSON_PATH.exists():
        return None
    try:
        payload = json.loads(CRM_JSON_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if payload.get("source_mtime_ns") != workbook_mtime_ns:
        return None
    if "rows" not in payload:
        return None
    if not any(key in payload for key in DETAIL_SHEETS):
        return None
    return payload


def read_crm_payload():
    stat = WORKBOOK_PATH.stat()
    if CRM_CACHE["payload"] is not None and CRM_CACHE["mtime_ns"] == stat.st_mtime_ns:
        return CRM_CACHE["payload"]
    source_payload = read_sidecar_payload(stat.st_mtime_ns) or read_workbook_payload()
    rows = source_payload.get("rows", [])
    updated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(stat.st_mtime))
    payload = {
        "rows": rows,
        "ratios": source_payload.get("ratios", []),
        "qa": source_payload.get("qa", []),
        "documents": source_payload.get("documents", []),
        "concepts": source_payload.get("concepts", []),
        "loanTape": source_payload.get("loanTape", []),
        "financialStatements": source_payload.get("financialStatements", []),
        "summary": summarize(rows, source_payload),
        "workbookUpdated": updated,
        "version": APP_VERSION,
    }
    CRM_CACHE.update({"mtime_ns": stat.st_mtime_ns, "payload": payload})
    return payload


def summarize(rows, payload=None):
    payload = payload or {}
    net_portfolio = 0
    loan_tape_review = 0
    for row in rows:
        try:
            net_portfolio += float(row.get("Cartera neta") or 0)
        except (TypeError, ValueError):
            pass
        if row.get("Calidad cartera") in {"Revisar", "Sin cartera"}:
            loan_tape_review += 1
    ratios = payload.get("ratios", [])
    qa = payload.get("qa", [])
    documents = payload.get("documents", [])
    concepts = payload.get("concepts", [])
    financial_statements = payload.get("financialStatements", [])
    calculated_ratios = sum(1 for row in ratios if row.get("review_status") == "calculated")
    reviewed_ratios = sum(1 for row in ratios if row.get("review_status") == "needs_review")
    ok_qa = sum(1 for row in qa if row.get("status") == "ok")
    pending_qa = sum(1 for row in qa if row.get("status") == "needs_review")
    return {
        "clients": len(rows),
        "high": sum(1 for row in rows if row.get("Prioridad") == "Alta"),
        "netPortfolio": net_portfolio,
        "loanTapeReview": loan_tape_review,
        "ratioReview": reviewed_ratios or sum(int(row.get("Razones revisar") or 0) for row in rows),
        "qaReview": pending_qa or sum(int(row.get("QA revisar") or 0) for row in rows),
        "documents": len(documents),
        "concepts": len(concepts),
        "financialStatementPeriods": len(
            {
                key
                for row in financial_statements
                for key, value in row.items()
                if isinstance(key, str)
                and len(key) == 10
                and key[:2] == "20"
                and value not in ("", None)
            }
        ),
        "ratios": len(ratios),
        "calculatedRatios": calculated_ratios,
        "qaOk": ok_qa,
    }


def load_followups():
    rows = []
    if FOLLOWUPS_PATH.exists():
        with FOLLOWUPS_PATH.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            for row in reader:
                rows.append({column: row.get(column, "") for column in FOLLOWUP_COLUMNS})
    return rows


def save_followup(update):
    client = str(update.get("client", "")).strip()
    if not client:
        raise ValueError("Missing client")
    rows = load_followups()
    by_client = {row["client"]: row for row in rows if row.get("client")}
    current = by_client.get(client, {"client": client})
    for column in FOLLOWUP_COLUMNS[1:]:
        current[column] = str(update.get(column, "") or "")
    by_client[client] = current
    ordered_clients = [row["client"] for row in rows if row.get("client")]
    if client not in ordered_clients:
        ordered_clients.append(client)
    FOLLOWUPS_PATH.parent.mkdir(parents=True, exist_ok=True)
    temp_path = FOLLOWUPS_PATH.with_suffix(".tmp")
    with temp_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FOLLOWUP_COLUMNS, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        for item in ordered_clients:
            writer.writerow({column: by_client[item].get(column, "") for column in FOLLOWUP_COLUMNS})
    temp_path.replace(FOLLOWUPS_PATH)


class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        return

    def do_GET(self):
        route = urlparse(self.path).path
        if route == "/":
            body = HTML.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Cache-Control", "no-store")
            self.send_header("X-CRM-Version", APP_VERSION)
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if route == "/api/crm":
            try:
                _json_response(self, 200, read_crm_payload())
            except Exception as exc:
                _json_response(self, 500, {"error": str(exc)})
            return
        _json_response(self, 404, {"error": "Not found"})

    def do_POST(self):
        route = urlparse(self.path).path
        if route == "/api/followups":
            try:
                save_followup(_read_body(self))
                _json_response(self, 200, {"ok": True})
            except Exception as exc:
                _json_response(self, 400, {"error": str(exc)})
            return
        if route == "/api/run":
            try:
                result = subprocess.run(
                    [str(RUN_SCRIPT)],
                    cwd=str(ROOT),
                    text=True,
                    capture_output=True,
                    timeout=900,
                    check=False,
                )
                if result.returncode:
                    _json_response(self, 500, {"error": result.stderr or result.stdout or "Pipeline failed"})
                else:
                    CRM_CACHE.update({"mtime_ns": None, "payload": None})
                    lines = [line for line in result.stdout.splitlines() if line.strip()]
                    _json_response(self, 200, {"message": lines[-1] if lines else "Procesado."})
            except Exception as exc:
                _json_response(self, 500, {"error": str(exc)})
            return
        if route == "/api/open-workbook":
            try:
                subprocess.Popen(["open", str(WORKBOOK_PATH)])
                _json_response(self, 200, {"ok": True})
            except Exception as exc:
                _json_response(self, 500, {"error": str(exc)})
            return
        _json_response(self, 404, {"error": "Not found"})


def main():
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8765
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"http://127.0.0.1:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
