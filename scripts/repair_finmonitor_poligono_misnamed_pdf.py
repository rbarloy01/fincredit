from pathlib import Path

from openpyxl import load_workbook

from build_finmonitor_5_company_workbook import (
    CLIENTS,
    OUTPUT,
    best_catalog_match,
    catalog_rows,
    extract_pdf,
    style_sheet,
)


TARGET_NAME = "ESTADO DE RESULTADOS PC ABRIL 2025..xlsx"


def main():
    catalog = catalog_rows()
    client = "POLIGONO CAPITAL"
    root = CLIENTS[client]
    target = next(path for path in root.rglob(TARGET_NAME))
    match = best_catalog_match(target, [row for row in catalog if row["client"] == client])
    replacement_rows = extract_pdf(client, target, match)

    wb = load_workbook(OUTPUT)
    ws = wb[client]
    header = [cell.value for cell in ws[1]]
    source_file_col = header.index("source_file") + 1
    rows_to_delete = [
        row_idx
        for row_idx in range(2, ws.max_row + 1)
        if ws.cell(row_idx, source_file_col).value == TARGET_NAME
    ]
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)
    for row in replacement_rows:
        ws.append(row)
    style_sheet(ws)
    wb.save(OUTPUT)
    print(f"replaced_rows={len(rows_to_delete)} added_rows={len(replacement_rows)}")


if __name__ == "__main__":
    main()
