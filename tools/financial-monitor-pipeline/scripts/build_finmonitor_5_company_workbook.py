import re
import signal
import unicodedata
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


WORKSPACE = Path("/Users/syscap/Documents/New project 2")
CATALOG = Path("/Users/syscap/Downloads/Catalogo Estados Financieros Completos.xlsx")
CLIENTS_ROOT = Path(
    "/Users/syscap/Library/CloudStorage/GoogleDrive-rbarron@syscap.com.mx/Shared drives/"
    "Axcess - Crédito y Riesgo/1. Clientes/2. Activos & Prospectos"
)
OUTPUT = WORKSPACE / "outputs/finmonitor_catalog_ingest/finmonitor_5_companies_eeff.xlsx"

CLIENTS = {
    "VENTUS": CLIENTS_ROOT / "Ventus",
    "AFIX": CLIENTS_ROOT / "Afix",
    "COFINE": CLIENTS_ROOT / "COFINE",
    "CAPITAL X": CLIENTS_ROOT / "Capital X",
    "POLIGONO CAPITAL": CLIENTS_ROOT / "Poligono Capital",
}

HEADERS = [
    "catalog_client",
    "period",
    "statement_type",
    "source_format",
    "source_file",
    "source_location",
    "row_type",
    "page_or_sheet",
    "line_or_row",
    "label",
    "value_1",
    "value_2",
    "value_3",
    "value_4",
    "value_5",
    "value_6",
    "raw_text",
    "local_path",
    "catalog_title_match",
    "catalog_link",
]

MONEY_RE = re.compile(r"\(?-?\s*\d{1,3}(?:[,\s]\d{3})+(?:\.\d+)?\)?|\(?-?\s*\d+(?:\.\d+)?\)?")
PAGE_TIMEOUT_SECONDS = 18


class PageTimeout(Exception):
    pass


def _timeout_handler(signum, frame):
    raise PageTimeout(f"page extraction timed out after {PAGE_TIMEOUT_SECONDS}s")


def normalize(text):
    text = "" if text is None else str(text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "—"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace(",", "").replace(" ", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return -number if negative else number


def period_from_name(name):
    match = re.search(r"(\d{6})", name)
    if match:
        token = match.group(1)
        return f"20{token[:2]}-{token[2:4]}-{token[4:6]}"
    match = re.search(r"\b(20\d{2})\b", name)
    return f"{match.group(1)}-12-31" if match else ""


def statement_type(name):
    norm = normalize(name)
    if re.search(r"\bbg\b|balance general|situacion financiera", norm):
        return "BG"
    if re.search(r"\ber\b|estado de resultados|resultados", norm):
        return "ER"
    if re.search(r"\bef\b|eeff|estado financiero|estados financieros", norm):
        return "EF"
    return "EEFF"


def catalog_rows():
    wb = load_workbook(CATALOG, read_only=False, data_only=True)
    ws = wb["Catálogo EEFF"]
    rows = []
    for row in range(2, ws.max_row + 1):
        client = ws.cell(row, 2).value
        title = ws.cell(row, 3).value
        fmt = ws.cell(row, 4).value
        if not client or not title:
            continue
        link_cell = ws.cell(row, 5)
        rows.append(
            {
                "client": str(client).strip().upper(),
                "title": str(title).strip(),
                "format": str(fmt or "").strip(),
                "link": link_cell.hyperlink.target if link_cell.hyperlink else str(link_cell.value or ""),
                "norm": normalize(title),
            }
        )
    return rows


def find_financial_files(client_root):
    suffixes = {".pdf", ".xlsx", ".xlsm", ".xls"}
    files = []
    for path in client_root.rglob("*"):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in suffixes:
            continue
        norm_path = normalize(path)
        if "informacion financiera" not in norm_path or "estados financieros" not in norm_path:
            continue
        files.append(path)
    return sorted(files, key=lambda p: normalize(p.name))


def is_pdf_file(path):
    try:
        with path.open("rb") as handle:
            return handle.read(5) == b"%PDF-"
    except OSError:
        return path.suffix.lower() == ".pdf"


def best_catalog_match(path, client_catalog):
    file_norm = normalize(path.stem)
    best = None
    best_score = 0
    file_tokens = set(file_norm.split())
    for row in client_catalog:
        title_tokens = set(row["norm"].split())
        score = len(file_tokens & title_tokens)
        if row["norm"] in file_norm or file_norm in row["norm"]:
            score += 20
        if period_from_name(row["title"]) and period_from_name(row["title"]) == period_from_name(path.name):
            score += 5
        if score > best_score:
            best = row
            best_score = score
    return best if best_score >= 3 else None


def values_from_cells(cells):
    values = []
    labels = []
    for cell in cells:
        if cell is None:
            continue
        number = clean_number(cell)
        if number is None:
            labels.append(str(cell).strip())
        else:
            values.append(number)
    label = " ".join(part for part in labels if part)
    return label, values[:6]


def row_payload(client, path, row_type, location, line_no, label, values, raw_text, match):
    vals = values[:6] + [None] * (6 - len(values))
    return [
        client,
        period_from_name(path.name),
        statement_type(path.name),
        path.suffix.upper().lstrip("."),
        path.name,
        str(location),
        row_type,
        location,
        line_no,
        label[:300],
        *vals,
        raw_text[:1000],
        str(path),
        match["title"] if match else "",
        match["link"] if match else "",
    ]


def extract_pdf(client, path, match):
    rows = []
    try:
        with pdfplumber.open(path) as pdf:
            for page_no, page in enumerate(pdf.pages, start=1):
                signal.signal(signal.SIGALRM, _timeout_handler)
                signal.alarm(PAGE_TIMEOUT_SECONDS)
                try:
                    text = page.extract_text() or ""
                except PageTimeout as exc:
                    rows.append(row_payload(client, path, "extract_error", f"p{page_no}", "", "ERROR", [], str(exc), match))
                    continue
                finally:
                    signal.alarm(0)
                for line_no, line in enumerate(text.splitlines(), start=1):
                    values = [clean_number(v) for v in MONEY_RE.findall(line)]
                    values = [v for v in values if v is not None]
                    if not values:
                        continue
                    label = MONEY_RE.sub(" ", line)
                    label = re.sub(r"\s+", " ", label).strip()
                    rows.append(row_payload(client, path, "pdf_text_line", f"p{page_no}", line_no, label, values, line, match))
                signal.alarm(PAGE_TIMEOUT_SECONDS)
                try:
                    tables = page.extract_tables() or []
                except PageTimeout as exc:
                    rows.append(row_payload(client, path, "extract_error", f"p{page_no}", "", "ERROR", [], str(exc), match))
                    tables = []
                except Exception:
                    tables = []
                finally:
                    signal.alarm(0)
                for table_no, table in enumerate(tables, start=1):
                    for table_row_no, table_row in enumerate(table or [], start=1):
                        label, values = values_from_cells(table_row)
                        if values:
                            raw = " | ".join("" if v is None else str(v) for v in table_row)
                            rows.append(
                                row_payload(
                                    client,
                                    path,
                                    "pdf_table_row",
                                    f"p{page_no}:t{table_no}",
                                    table_row_no,
                                    label,
                                    values,
                                    raw,
                                    match,
                                )
                            )
    except Exception as exc:
        rows.append(row_payload(client, path, "extract_error", "", "", "ERROR", [], str(exc), match))
    return rows


def extract_workbook(client, path, match):
    rows = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
                label, values = values_from_cells(row)
                if not values:
                    continue
                raw = " | ".join("" if v is None else str(v) for v in row)
                rows.append(row_payload(client, path, "xlsx_row", ws.title[:31], row_no, label, values, raw, match))
    except Exception as exc:
        rows.append(row_payload(client, path, "extract_error", "", "", "ERROR", [], str(exc), match))
    return rows


def autosize(ws):
    widths = {
        "A": 18,
        "B": 12,
        "C": 10,
        "D": 10,
        "E": 34,
        "F": 16,
        "J": 42,
        "Q": 64,
        "R": 70,
        "S": 34,
        "T": 48,
    }
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(letter, 14 if 11 <= col <= 16 else 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical="top")
            if 11 <= cell.column <= 16:
                cell.number_format = '#,##0.00;[Red](#,##0.00);-'
    autosize(ws)


def main():
    catalog = catalog_rows()
    wb = Workbook()
    wb.remove(wb.active)

    for client, root in CLIENTS.items():
        ws = wb.create_sheet(client[:31])
        ws.append(HEADERS)
        client_catalog = [row for row in catalog if row["client"] == client]
        files = find_financial_files(root)
        print(f"{client}: files={len(files)}", flush=True)
        for path in files:
            print(f"  extracting {path.name}", flush=True)
            match = best_catalog_match(path, client_catalog)
            if is_pdf_file(path):
                rows = extract_pdf(client, path, match)
            else:
                rows = extract_workbook(client, path, match)
            for row in rows:
                ws.append(row)
        style_sheet(ws)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT)
    for ws in wb.worksheets:
        print(f"{ws.title}: rows={ws.max_row - 1}")


if __name__ == "__main__":
    main()
